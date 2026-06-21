from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QTableWidget,
    QTableWidgetItem, QHeaderView,
    QAbstractItemView, QMessageBox,
    QFileDialog, QButtonGroup, QRadioButton,
    QFrame
)
from PySide6.QtCore import Qt
from datetime import datetime
import os

BTN_STYLE = """
    QPushButton {
        background:#F5F5F5; border:1px solid #DDD;
        border-radius:5px; padding:5px 12px;
        font-size:12px; color:#333;
    }
    QPushButton:hover { background:#E8F5F0; border-color:#1D9E75; color:#1D9E75; }
"""
BTN_PRIMARY = """
    QPushButton {
        background:#1D9E75; color:white;
        border-radius:5px; padding:5px 14px;
        font-size:12px; font-weight:600; border:none;
    }
    QPushButton:hover { background:#0F6E56; }
    QPushButton:disabled { background:#A5D6C5; }
"""
BTN_DANGER = """
    QPushButton {
        background:#E53935; color:white;
        border-radius:5px; padding:5px 14px;
        font-size:12px; font-weight:600; border:none;
    }
    QPushButton:hover { background:#B71C1C; }
    QPushButton:disabled { background:#FFCDD2; }
"""

MODE_MERGE   = "merge"    # Cập nhật thông tin mới
MODE_REPLACE = "replace"  # Xóa cũ, thêm mới hoàn toàn


class ImportExcelDialog(QDialog):
    def __init__(self, parent=None, ten_lop="", existing_count=0):
        """
        existing_count: số học sinh hiện có trong lớp,
        dùng để hiển thị cảnh báo ở chế độ Replace.
        """
        super().__init__(parent)
        self.setWindowTitle(f"📥 Import học sinh từ Excel — Lớp {ten_lop}")
        self.setMinimumSize(720, 560)
        self._rows = []
        self._existing_count = existing_count
        self._build_ui()

    # ──────────────────────────────────────────────
    # BUILD UI
    # ──────────────────────────────────────────────
    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(15, 15, 15, 15)

        # Hướng dẫn
        lbl_hd = QLabel(
            "📌 Hướng dẫn: Tải file mẫu → Điền thông tin → Chọn file → Kiểm tra → Import"
        )
        lbl_hd.setStyleSheet(
            "background:#E8F5E9; border:1px solid #A5D6C5; "
            "border-radius:5px; padding:8px; font-size:12px; color:#2E7D32;"
        )
        layout.addWidget(lbl_hd)

        # ── Chế độ import ──
        layout.addWidget(self._build_mode_section())

        # Toolbar chọn file
        tb = QHBoxLayout()
        tb.setSpacing(8)

        btn_mau = QPushButton("⬇️ Tải file mẫu")
        btn_mau.setStyleSheet(BTN_STYLE)
        btn_mau.setFixedHeight(32)
        btn_mau.clicked.connect(self._download_template)

        btn_chon = QPushButton("📂 Chọn file Excel")
        btn_chon.setStyleSheet(BTN_STYLE)
        btn_chon.setFixedHeight(32)
        btn_chon.clicked.connect(self._pick_file)

        self.lbl_file = QLabel("Chưa chọn file")
        self.lbl_file.setStyleSheet("color:#888; font-size:12px;")

        tb.addWidget(btn_mau)
        tb.addWidget(btn_chon)
        tb.addWidget(self.lbl_file)
        tb.addStretch()
        layout.addLayout(tb)

        # Bảng preview
        lbl_preview = QLabel("👁 Xem trước dữ liệu:")
        lbl_preview.setStyleSheet("font-weight:600; font-size:12px;")
        layout.addWidget(lbl_preview)

        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels([
            "STT", "Họ và tên", "Ngày sinh", "Giới tính", "SĐT"
        ])
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.setColumnWidth(0, 40)
        self.table.setColumnWidth(2, 90)
        self.table.setColumnWidth(3, 70)
        self.table.setColumnWidth(4, 100)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget {
                border:1px solid #DDD;
                gridline-color:#E8E8E8;
                font-size:12px;
            }
            QHeaderView::section {
                background:#1D9E75; color:white;
                font-weight:600; padding:5px; border:none;
            }
            QTableWidget::item:alternate { background:#F9F9F9; }
        """)
        layout.addWidget(self.table)

        # Status
        self.lbl_status = QLabel("")
        self.lbl_status.setStyleSheet("font-size:12px; color:#555;")
        layout.addWidget(self.lbl_status)

        # Cảnh báo replace (ẩn mặc định)
        self.lbl_warning = QLabel("")
        self.lbl_warning.setStyleSheet(
            "background:#FFF3E0; border:1px solid #FFB74D; "
            "border-radius:5px; padding:7px; font-size:12px; color:#E65100;"
        )
        self.lbl_warning.setVisible(False)
        layout.addWidget(self.lbl_warning)

        # Nút OK/Cancel
        btn_row = QHBoxLayout()
        self.btn_import = QPushButton("✅ Import")
        self.btn_import.setStyleSheet(BTN_PRIMARY)
        self.btn_import.setFixedHeight(32)
        self.btn_import.setEnabled(False)
        self.btn_import.clicked.connect(self._on_import_clicked)

        btn_cancel = QPushButton("❌ Hủy")
        btn_cancel.setStyleSheet(BTN_STYLE)
        btn_cancel.setFixedHeight(32)
        btn_cancel.clicked.connect(self.reject)

        btn_row.addStretch()
        btn_row.addWidget(self.btn_import)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

    def _build_mode_section(self) -> QFrame:
        """Khung chọn chế độ import."""
        frame = QFrame()
        frame.setStyleSheet(
            "QFrame { background:#F8F8F8; border:1px solid #DDD; border-radius:6px; }"
        )
        h = QHBoxLayout(frame)
        h.setContentsMargins(12, 8, 12, 8)
        h.setSpacing(20)

        lbl = QLabel("Chế độ import:")
        lbl.setStyleSheet("font-weight:600; font-size:12px; border:none; background:transparent;")
        h.addWidget(lbl)

        self._mode_group = QButtonGroup(self)

        self.rb_merge = QRadioButton("🔄  Cập nhật thông tin mới")
        self.rb_merge.setStyleSheet("font-size:12px; border:none; background:transparent;")
        self.rb_merge.setChecked(True)
        self.rb_merge.setToolTip(
            "Học sinh đã có (khớp Họ tên + Ngày sinh) → cập nhật.\n"
            "Học sinh mới → thêm vào.\n"
            "Học sinh cũ không có trong file → giữ nguyên."
        )

        self.rb_replace = QRadioButton("🗑️  Xóa cũ, thêm mới hoàn toàn")
        self.rb_replace.setStyleSheet("font-size:12px; border:none; background:transparent;")
        self.rb_replace.setToolTip(
            "Toàn bộ danh sách học sinh hiện tại sẽ bị xóa.\n"
            "Danh sách mới được thêm từ file Excel."
        )

        self._mode_group.addButton(self.rb_merge,   0)
        self._mode_group.addButton(self.rb_replace, 1)

        h.addWidget(self.rb_merge)
        h.addWidget(self.rb_replace)
        h.addStretch()

        # Cập nhật UI khi đổi mode
        self.rb_merge.toggled.connect(self._on_mode_changed)
        self.rb_replace.toggled.connect(self._on_mode_changed)

        return frame

    # ──────────────────────────────────────────────
    # EVENTS
    # ──────────────────────────────────────────────
    def _on_mode_changed(self):
        self._refresh_status()
        self._refresh_warning()
        self._refresh_import_btn()

    def _on_import_clicked(self):
        if self.get_mode() == MODE_REPLACE and self._existing_count > 0:
            reply = QMessageBox.warning(
                self,
                "⚠️ Xác nhận xóa dữ liệu cũ",
                f"Thao tác này sẽ <b>xóa vĩnh viễn {self._existing_count} học sinh</b> "
                f"hiện có trong lớp và thay bằng {len(self._rows)} học sinh từ file.<br><br>"
                f"Bạn có chắc chắn muốn tiếp tục không?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if reply != QMessageBox.Yes:
                return
        self.accept()

    # ──────────────────────────────────────────────
    # HELPERS: refresh UI
    # ──────────────────────────────────────────────
    def _refresh_status(self):
        if not self._rows:
            return
        n = len(self._rows)
        if self.get_mode() == MODE_MERGE:
            self.lbl_status.setText(
                f"✅ Đọc được {n} học sinh — "
                "sẽ cập nhật trùng / thêm mới, giữ nguyên phần còn lại."
            )
        else:
            self.lbl_status.setText(
                f"✅ Đọc được {n} học sinh — "
                f"sẽ xóa toàn bộ danh sách cũ và thay thế hoàn toàn."
            )

    def _refresh_warning(self):
        if self.get_mode() == MODE_REPLACE and self._existing_count > 0 and self._rows:
            self.lbl_warning.setText(
                f"⚠️  Chế độ <b>Xóa & Thay mới</b>: "
                f"{self._existing_count} học sinh hiện tại sẽ bị xóa, "
                f"thay bằng {len(self._rows)} học sinh từ file."
            )
            self.lbl_warning.setVisible(True)
        else:
            self.lbl_warning.setVisible(False)

    def _refresh_import_btn(self):
        has_rows = len(self._rows) > 0
        self.btn_import.setEnabled(has_rows)
        if has_rows and self.get_mode() == MODE_REPLACE:
            self.btn_import.setText("🗑️ Xóa cũ & Import")
            self.btn_import.setStyleSheet(BTN_DANGER)
        else:
            self.btn_import.setText("✅ Import")
            self.btn_import.setStyleSheet(BTN_PRIMARY)

    # ──────────────────────────────────────────────
    # DOWNLOAD TEMPLATE
    # ──────────────────────────────────────────────
    def _download_template(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Danh sách học sinh"

            header_fill = PatternFill("solid", fgColor="1D9E75")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_align = Alignment(horizontal="center", vertical="center")
            thin = Side(style="thin", color="CCCCCC")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            headers = ["STT", "Họ và tên", "Ngày sinh", "Giới tính", "Số điện thoại"]
            col_widths = [8, 30, 15, 12, 18]

            for col, (h, w) in enumerate(zip(headers, col_widths), 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = border
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col)
                ].width = w

            ws.row_dimensions[1].height = 25

            examples = [
                (1, "Nguyễn Văn An", "01/09/2012", "Nam", "0912345678"),
                (2, "Trần Thị Bình", "15/03/2012", "Nữ", "0987654321"),
            ]
            center = Alignment(horizontal="center", vertical="center")
            left   = Alignment(horizontal="left",   vertical="center")

            for row_idx, row_data in enumerate(examples, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = left if col_idx == 2 else center
                ws.row_dimensions[row_idx].height = 20

            save_path, _ = QFileDialog.getSaveFileName(
                self, "Lưu file mẫu", "mau_hoc_sinh.xlsx",
                "Excel Files (*.xlsx)"
            )
            if save_path:
                wb.save(save_path)
                QMessageBox.information(self, "Thành công",
                    f"✅ Đã lưu file mẫu!\n{save_path}")

        except ImportError:
            QMessageBox.warning(self, "Lỗi",
                "Cần cài thư viện openpyxl!\nChạy: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    # ──────────────────────────────────────────────
    # PICK & READ FILE
    # ──────────────────────────────────────────────
    def _pick_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Chọn file Excel", "",
            "Excel Files (*.xlsx *.xls)"
        )
        if not path:
            return
        self.lbl_file.setText(os.path.basename(path))
        self._read_excel(path)

    def _read_excel(self, path: str):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)

            if wb is None:
                QMessageBox.warning(self, "Lỗi", "Không mở được file Excel!")
                return

            ws = wb.active
            if ws is None:
                QMessageBox.warning(self, "Lỗi", "File Excel không có sheet nào!")
                return

            self._rows = []
            errors = []
            self.table.setRowCount(0)

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not any(row):
                    continue

                stt            = row[0]
                ho_ten         = str(row[1]).strip() if row[1] else ""
                ngay_sinh_raw  = row[2]
                gioi_tinh_raw  = str(row[3]).strip() if row[3] else ""
                so_dt          = str(row[4]).strip() if row[4] else ""

                if not ho_ten:
                    errors.append(f"Dòng {row_idx}: Thiếu họ tên")
                    continue

                ngay_sinh = None
                if ngay_sinh_raw:
                    if isinstance(ngay_sinh_raw, datetime):
                        ngay_sinh = ngay_sinh_raw.date()
                    else:
                        try:
                            ngay_sinh = datetime.strptime(
                                str(ngay_sinh_raw).strip(), "%d/%m/%Y"
                            ).date()
                        except Exception:
                            errors.append(
                                f"Dòng {row_idx}: Ngày sinh sai định dạng (dd/mm/yyyy)"
                            )
                            continue

                gioi_tinh = True
                if gioi_tinh_raw.lower() in ("nữ", "nu", "female", "f", "0"):
                    gioi_tinh = False

                # Chuẩn hóa key so sánh: lowercase + bỏ khoảng trắng thừa
                _key = (
                    " ".join(ho_ten.lower().split()),
                    ngay_sinh,
                )

                self._rows.append({
                    "ho_ten":         ho_ten,
                    "ngay_sinh":      ngay_sinh,
                    "gioi_tinh":      gioi_tinh,
                    "so_dien_thoai":  so_dt or None,
                    "_match_key":     _key,   # dùng nội bộ để so khớp
                })

                i = self.table.rowCount()
                self.table.insertRow(i)
                self.table.setItem(i, 0, self._cell(str(stt or i + 1), Qt.AlignCenter))
                self.table.setItem(i, 1, self._cell(ho_ten))
                self.table.setItem(i, 2, self._cell(
                    ngay_sinh.strftime("%d/%m/%Y") if ngay_sinh else "", Qt.AlignCenter
                ))
                self.table.setItem(i, 3, self._cell(
                    "Nam" if gioi_tinh else "Nữ", Qt.AlignCenter
                ))
                self.table.setItem(i, 4, self._cell(so_dt, Qt.AlignCenter))

            # Cập nhật toàn bộ UI sau khi đọc xong
            status_base = f"✅ Đọc được {len(self._rows)} học sinh"
            if errors:
                status_base += (
                    f" | ⚠️ {len(errors)} lỗi: {'; '.join(errors[:3])}"
                    + (f"... (+{len(errors)-3})" if len(errors) > 3 else "")
                )
            self.lbl_status.setText(status_base)

            self._refresh_status()
            self._refresh_warning()
            self._refresh_import_btn()

        except ImportError:
            QMessageBox.warning(self, "Lỗi",
                "Cần cài thư viện openpyxl!\nChạy: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không đọc được file!\n{str(e)}")

    # ──────────────────────────────────────────────
    # CELL HELPER
    # ──────────────────────────────────────────────
    def _cell(self, text, align=Qt.AlignLeft | Qt.AlignVCenter):
        item = QTableWidgetItem(str(text))
        item.setTextAlignment(align)
        return item

    # ──────────────────────────────────────────────
    # PUBLIC API
    # ──────────────────────────────────────────────
    def get_rows(self):
        """Trả về danh sách học sinh đã đọc từ file."""
        return self._rows

    def get_mode(self) -> str:
        """Trả về MODE_MERGE hoặc MODE_REPLACE."""
        return MODE_REPLACE if self.rb_replace.isChecked() else MODE_MERGE