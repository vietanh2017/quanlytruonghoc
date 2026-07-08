import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from sqlalchemy.orm import Session

from modules.tkb.models import ThoiKhoaBieu
from core.db.models.phan_cong import PhanCongGiangDay
from core.db.models.nam_hoc import NamHoc
from core.db.models.hoc_ky import HocKy
from core.db.models.lop_hoc import LopHoc
from core.db.models.giao_vien import GiaoVien
from core.db.models.mon_hoc import MonHoc, MonHocKhoi


class TKBExportService:
    def __init__(self, session: Session):
        self.session = session
        
        # ═══ MÀU SẮC ═══
        self.COLOR_HEADER = "D9E1F2"      # Xanh nhạt header
        self.COLOR_SUB_HEADER = "E2EFDA"  # Xanh lá nhạt sub-header
        self.COLOR_ALTERNATE = "F5F5F5"   # Xám xen kẽ
        self.COLOR_WHITE = "FFFFFF"       # Trắng
        self.COLOR_BORDER = "D0D0D0"      # Màu viền
        
        # ═══ FONT ═══
        self.FONT_HEADER = Font(name='Times New Roman', size=13, bold=True)
        self.FONT_SUB_HEADER = Font(name='Times New Roman', size=12, bold=True)
        self.FONT_NORMAL = Font(name='Times New Roman', size=11)
        self.FONT_NORMAL_BOLD = Font(name='Times New Roman', size=11, bold=True)
        self.FONT_SMALL = Font(name='Times New Roman', size=10)
        self.FONT_SMALL_BOLD = Font(name='Times New Roman', size=10, bold=True)

    def export_tkb(self, nam_hoc_id: int, hoc_ky_id: int) -> io.BytesIO:
        """Xuất toàn bộ TKB ra file Excel"""
        wb = Workbook()
        
        # Lấy thông tin trường và năm học
        thong_tin = self._get_thong_tin_truong()
        nam_hoc_info = self._get_nam_hoc_hoc_ky(nam_hoc_id, hoc_ky_id)
        
        # Sheet 1: Phân công giảng dạy
        self._create_sheet_pcgd(wb, nam_hoc_id, hoc_ky_id, thong_tin, nam_hoc_info)
        
        # Sheet 2: TKB theo lớp
        self._create_sheet_tkb_lop(wb, nam_hoc_id, hoc_ky_id, thong_tin, nam_hoc_info)
        
        # Sheet 3: TKB theo giáo viên
        self._create_sheet_tkb_gv(wb, nam_hoc_id, hoc_ky_id, thong_tin, nam_hoc_info)
        
        # Xóa sheet mặc định
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _get_thong_tin_truong(self):
        """Lấy thông tin trường"""
        try:
            from modules.cau_hinh.repository import ThongTinTruong
            info = self.session.query(ThongTinTruong).first()
            if info:
                return {
                    'ten_truong': info.ten_truong or 'TRƯỜNG THCS ...',
                    'dia_chi': info.dia_chi or '',
                    'dien_thoai': info.dien_thoai or '',
                    'email': info.email or '',
                    'hieu_truong': info.hieu_truong or '',
                    'hieu_pho': info.hieu_pho or '',
                    'to_truong_cm': info.to_truong_cm or '',
                }
        except:
            pass
        return {
            'ten_truong': 'TRƯỜNG THCS ...',
            'dia_chi': '',
            'dien_thoai': '',
            'email': '',
            'hieu_truong': '',
            'hieu_pho': '',
            'to_truong_cm': '',
        }

    def _get_nam_hoc_hoc_ky(self, nam_hoc_id, hoc_ky_id):
        """Lấy thông tin năm học và học kỳ"""
        nam_hoc = self.session.query(NamHoc).filter(NamHoc.id == nam_hoc_id).first()
        hoc_ky = self.session.query(HocKy).filter(HocKy.id == hoc_ky_id).first()
        return {
            'ten_nam_hoc': nam_hoc.ten_nam_hoc if nam_hoc else '',
            'ten_hoc_ky': hoc_ky.ten_hoc_ky if hoc_ky else '',
        }

    def _apply_border(self, ws, start_row, start_col, end_row, end_col):
        """Áp dụng viền cho vùng dữ liệu"""
        border = Border(
            left=Side(style='thin', color=self.COLOR_BORDER),
            right=Side(style='thin', color=self.COLOR_BORDER),
            top=Side(style='thin', color=self.COLOR_BORDER),
            bottom=Side(style='thin', color=self.COLOR_BORDER)
        )
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = border

    # ════════════════════════════════════════════════════════════
    #  SHEET 1: PHÂN CÔNG GIẢNG DẠY (PCGD)
    # ════════════════════════════════════════════════════════════

    def _create_sheet_pcgd(self, wb, nam_hoc_id, hoc_ky_id, thong_tin, nam_hoc_info):
        ws = wb.create_sheet("PCGD")
        
        # ── HEADER ──────────────────────────────────────────
        row = 1
        
        # Dòng 1: Tên trường
        ws.merge_cells(f'A{row}:I{row}')  # ⭐ Tăng lên 9 cột
        ws[f'A{row}'].value = thong_tin['ten_truong'].upper()
        ws[f'A{row}'].font = Font(name='Times New Roman', size=14, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # Dòng 2: Năm học - Học kỳ
        ws.merge_cells(f'A{row}:I{row}')
        ws[f'A{row}'].value = f"Năm học {nam_hoc_info['ten_nam_hoc']} - {nam_hoc_info['ten_hoc_ky']}"
        ws[f'A{row}'].font = Font(name='Times New Roman', size=12)
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # Dòng 3: Tiêu đề bảng
        ws.merge_cells(f'A{row}:I{row}')
        ws[f'A{row}'].value = "BẢNG PHÂN CÔNG GIẢNG DẠY"
        ws[f'A{row}'].font = Font(name='Times New Roman', size=14, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        row += 2
        
        # ── HEADER BẢNG ────────────────────────────────────
        # ⭐ THÊM CÁC CỘT MỚI
        headers = [
            'TT', 'Giáo viên', 'Kiêm nhiệm', 'CN', 
            'Phân công chuyên môn', 'Thực dạy', 
            'Tiết CN', 'Tiết KN', 'Tổng'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(name='Times New Roman', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            self._apply_border_to_cell(ws, row, col)
        row += 1
        
        # ── DỮ LIỆU ────────────────────────────────────────
        phan_congs = self.session.query(PhanCongGiangDay).filter(
            PhanCongGiangDay.nam_hoc_id == nam_hoc_id,
            PhanCongGiangDay.hoc_ky_id == hoc_ky_id
        ).all()
        
        # Lấy danh sách lớp để biết GV chủ nhiệm
        ds_lop = self.session.query(LopHoc).all()
        gvcn_map = {}  # {gv_id: [ten_lop]}
        for lop in ds_lop:
            if lop.giao_vien_cn_id:
                if lop.giao_vien_cn_id not in gvcn_map:
                    gvcn_map[lop.giao_vien_cn_id] = []
                gvcn_map[lop.giao_vien_cn_id].append(lop.ten_lop)
        
        # ⭐ Import hàm lấy số tiết giảm (kiêm nhiệm)
        from modules.cau_hinh.kiem_nhiem_config import get_so_tiet_giam
        
        # ⭐ Số tiết chủ nhiệm theo quy định (Thông tư 28)
        SO_TIET_CN = 4  # 1 lớp chủ nhiệm = 4 tiết/tuần
        
        # Nhóm theo giáo viên
        gv_map = {}
        for pc in phan_congs:
            gv = pc.giao_vien
            if not gv:
                continue
            gv_id = gv.id
            if gv_id not in gv_map:
                gv_map[gv_id] = {
                    'gv': gv,
                    'mon_lop': [],
                    'tong_tiet_day': 0  # ⭐ Số tiết thực dạy
                }
            mon = pc.mon_hoc
            lop = pc.lop_hoc
            if mon and lop:
                ten_mon = self._get_ten_mon_hien_thi(
                    phan_cong_id=pc.id,
                    mon_hoc=mon,
                    lop_hoc=lop
                )
                mon_lop = f"{ten_mon} ({lop.ten_lop})"
                gv_map[gv_id]['mon_lop'].append(mon_lop)
                mk = self.session.query(MonHocKhoi).filter(
                    MonHocKhoi.mon_hoc_id == mon.id,
                    MonHocKhoi.khoi == lop.khoi
                ).first()
                so_tiet = mk.so_tiet if mk else 1
                gv_map[gv_id]['tong_tiet_day'] += so_tiet
        
        # Xuất dữ liệu
        stt = 1
        start_row = row
        for gv_id, data in gv_map.items():
            gv = data['gv']
            ten_gv = gv.nguoi_dung.ho_ten if gv.nguoi_dung else ''
            
            # Lấy KIÊM NHIỆM
            kiem_nhiem = getattr(gv, 'kiem_nhiem', '') or ''
            
            # Lấy CHỦ NHIỆM (danh sách lớp)
            cn_list = gvcn_map.get(gv_id, [])
            cn_str = ', '.join(cn_list) if cn_list else ''
            
            mon_lop_str = ', '.join(data['mon_lop'])
            so_tiet_day = data['tong_tiet_day']  # Số tiết thực dạy
            
            # ⭐ TÍNH TIẾT CHỦ NHIỆM (mỗi lớp CN = 4 tiết)
            so_tiet_cn = len(cn_list) * SO_TIET_CN
            
            # ⭐ TÍNH TIẾT KIÊM NHIỆM
            so_tiet_kn = get_so_tiet_giam(kiem_nhiem)  # Số tiết được cộng thêm
            
            # ⭐ TỔNG SỐ TIẾT
            tong = so_tiet_day + so_tiet_cn + so_tiet_kn
            
            # Xen kẽ màu
            bg_color = self.COLOR_ALTERNATE if stt % 2 == 0 else self.COLOR_WHITE
            
            ws.cell(row=row, column=1, value=stt).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=2, value=ten_gv)
            ws.cell(row=row, column=3, value=kiem_nhiem)
            ws.cell(row=row, column=4, value=cn_str)
            ws.cell(row=row, column=5, value=mon_lop_str)
            ws.cell(row=row, column=6, value=so_tiet_day).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=7, value=so_tiet_cn if so_tiet_cn > 0 else '').alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=8, value=so_tiet_kn if so_tiet_kn > 0 else '').alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=9, value=tong).alignment = Alignment(horizontal='center', vertical='center')
            
            # Áp dụng màu xen kẽ và viền
            for col in range(1, 10):  # ⭐ 9 cột
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=bg_color, end_color=bg_color, fill_type="solid"
                )
                self._apply_border_to_cell(ws, row, col)
            
            stt += 1
            row += 1
        
        # ── VIỀN ────────────────────────────────────────────
        if stt > 1:
            self._apply_border(ws, start_row - 1, 1, row - 1, 9)  # ⭐ 9 cột
        
        # ── ĐỘ RỘNG CỘT ────────────────────────────────────
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 12   # Số tiết dạy
        ws.column_dimensions['G'].width = 10   # Tiết CN
        ws.column_dimensions['H'].width = 10   # Tiết KN
        ws.column_dimensions['I'].width = 12   # Tổng

    # ════════════════════════════════════════════════════════════
    #  SHEET 2: TKB THEO LỚP
    # ════════════════════════════════════════════════════════════

    def _create_sheet_tkb_lop(self, wb, nam_hoc_id, hoc_ky_id, thong_tin, nam_hoc_info):
        ws = wb.create_sheet("TKB_Lớp")
        
        # ── LẤY DANH SÁCH LỚP ─────────────────────────────
        tkb_records = self.session.query(ThoiKhoaBieu).filter(
            ThoiKhoaBieu.nam_hoc_id == nam_hoc_id,
            ThoiKhoaBieu.hoc_ky_id == hoc_ky_id,
            ThoiKhoaBieu.is_active == True
        ).all()
        
        lop_ids = sorted(list(set([t.lop_hoc_id for t in tkb_records])))
        lop_hocs = []
        for lop_id in lop_ids:
            lop = self.session.query(LopHoc).filter(LopHoc.id == lop_id).first()
            if lop:
                lop_hocs.append(lop)
        
        if not lop_hocs:
            ws['A1'] = "Không có dữ liệu TKB"
            return
        
        # ── HEADER ──────────────────────────────────────────
        row = 1
        total_cols = 2 + len(lop_hocs) * 2
        last_col = get_column_letter(total_cols)
        
        # Dòng 1: Tên trường
        ws.merge_cells(f'A{row}:{last_col}{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = thong_tin['ten_truong'].upper()
        cell.font = Font(name='Times New Roman', size=14, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # ⭐ Viền cho header
        for c in range(1, total_cols + 1):
            self._apply_border_to_cell(ws, row, c)
        row += 1
        
        # Dòng 2: Năm học - Học kỳ
        ws.merge_cells(f'A{row}:{last_col}{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = f"Năm học {nam_hoc_info['ten_nam_hoc']} - {nam_hoc_info['ten_hoc_ky']}"
        cell.font = Font(name='Times New Roman', size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        for c in range(1, total_cols + 1):
            self._apply_border_to_cell(ws, row, c)
        row += 1
        
        # Dòng 3: Tiêu đề
        ws.merge_cells(f'A{row}:{last_col}{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = "THỜI KHÓA BIỂU"
        cell.font = Font(name='Times New Roman', size=14, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        for c in range(1, total_cols + 1):
            self._apply_border_to_cell(ws, row, c)
        row += 2
        
        # ── HEADER BẢNG ────────────────────────────────────
        header_row = row
        
        # Cột THỨ - merge 2 dòng
        ws.merge_cells(
            start_row=row, 
            start_column=1, 
            end_row=row + 1, 
            end_column=1
        )
        cell = ws.cell(row=row, column=1)
        cell.value = 'THỨ'
        cell.font = Font(name='Times New Roman', size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        self._apply_border_to_cell(ws, row, 1)
        
        # Cột TIẾT - merge 2 dòng
        ws.merge_cells(
            start_row=row, 
            start_column=2, 
            end_row=row + 1, 
            end_column=2
        )
        cell = ws.cell(row=row, column=2)
        cell.value = 'TIẾT'
        cell.font = Font(name='Times New Roman', size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        self._apply_border_to_cell(ws, row, 2)
        
        # Các cột lớp (mỗi lớp merge 2 cột Sáng + Chiều)
        lop_cols = {}
        for idx, lop in enumerate(lop_hocs):
            col_start = 2 + idx * 2 + 1
            col_end = col_start + 1
            lop_cols[lop.id] = (col_start, col_end)
            
            ws.merge_cells(
                start_row=row, 
                start_column=col_start, 
                end_row=row, 
                end_column=col_end
            )
            cell = ws.cell(row=row, column=col_start)
            
            ten_gvcn = ''
            if lop.giao_vien_cn_id:
                gvcn = self.session.query(GiaoVien).filter(
                    GiaoVien.id == lop.giao_vien_cn_id
                ).first()
                if gvcn and gvcn.nguoi_dung:
                    ten_gvcn = f"({gvcn.nguoi_dung.ho_ten})"
            
            cell.value = f"{lop.ten_lop} {ten_gvcn}".strip()
            cell.font = Font(name='Times New Roman', size=10, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            self._apply_border_to_cell(ws, row, col_start)
        
        row += 1
        
        # Hàng 2: Sáng | Chiều cho từng lớp
        cell = ws.cell(row=row, column=1)
        cell.fill = PatternFill(start_color=self.COLOR_SUB_HEADER, end_color=self.COLOR_SUB_HEADER, fill_type="solid")
        self._apply_border_to_cell(ws, row, 1)
        
        cell = ws.cell(row=row, column=2)
        cell.fill = PatternFill(start_color=self.COLOR_SUB_HEADER, end_color=self.COLOR_SUB_HEADER, fill_type="solid")
        self._apply_border_to_cell(ws, row, 2)
        
        for idx, lop in enumerate(lop_hocs):
            col_start, col_end = lop_cols[lop.id]
            
            cell = ws.cell(row=row, column=col_start)
            cell.value = 'Sáng'
            cell.font = Font(name='Times New Roman', size=9, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color=self.COLOR_SUB_HEADER, end_color=self.COLOR_SUB_HEADER, fill_type="solid")
            self._apply_border_to_cell(ws, row, col_start)
            
            cell = ws.cell(row=row, column=col_end)
            cell.value = 'Chiều'
            cell.font = Font(name='Times New Roman', size=9, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color=self.COLOR_SUB_HEADER, end_color=self.COLOR_SUB_HEADER, fill_type="solid")
            self._apply_border_to_cell(ws, row, col_end)
        
        row += 1
        
        # ── DỮ LIỆU ────────────────────────────────────────
        tkb_map = {}
        for t in tkb_records:
            key = (t.lop_hoc_id, t.thu, t.tiet, t.buoi)
            tkb_map[key] = t
        
        thu_list = [2, 3, 4, 5, 6, 7]
        thu_names = {2: '2', 3: '3', 4: '4', 5: '5', 6: '6', 7: '7'}
        max_tiet = 5
        start_row = row
        
        for thu in thu_list:
            for tiet in range(1, max_tiet + 1):
                # Cột THỨ - merge 5 dòng
                if tiet == 1:
                    ws.merge_cells(
                        start_row=row, 
                        start_column=1, 
                        end_row=row + max_tiet - 1, 
                        end_column=1
                    )
                    cell = ws.cell(row=row, column=1)
                    cell.value = thu_names[thu]
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(name='Times New Roman', size=10, bold=True)
                    self._apply_border_to_cell(ws, row, 1)
                    self._apply_border_to_cell(ws, row + max_tiet - 1, 1)
                
                # Cột TIẾT
                cell = ws.cell(row=row, column=2)
                cell.value = f"{tiet}"
                cell.alignment = Alignment(horizontal='center', vertical='center')
                self._apply_border_to_cell(ws, row, 2)
                
                # Dữ liệu từng lớp
                for lop in lop_hocs:
                    col_start, col_end = lop_cols[lop.id]
                    
                    # Buổi Sáng
                    key_sang = (lop.id, thu, tiet, 'sang')
                    tkb_sang = tkb_map.get(key_sang)
                    value = self._ten_mon(tkb_sang)
                    cell = ws.cell(row=row, column=col_start)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    self._apply_border_to_cell(ws, row, col_start)
                    
                    # Buổi Chiều
                    key_chieu = (lop.id, thu, tiet, 'chieu')
                    tkb_chieu = tkb_map.get(key_chieu)
                    value = self._ten_mon(tkb_chieu)
                    cell = ws.cell(row=row, column=col_end)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    self._apply_border_to_cell(ws, row, col_end)
                
                row += 1
        
        # ── VIỀN CHO TOÀN BỘ BẢNG ─────────────────────────
        # Áp dụng viền cho toàn bộ vùng từ header đến hết dữ liệu
        self._apply_border(ws, header_row, 1, row - 1, total_cols)
        
        # ── ĐỘ RỘNG CỘT ────────────────────────────────────
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 10
        for i in range(3, total_cols + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18
        
        # ── CHIỀU CAO HÀNG ──────────────────────────────────
        for r in range(header_row, row):
            ws.row_dimensions[r].height = 20

    # ════════════════════════════════════════════════════════════
    #  SHEET 3: TKB THEO GIÁO VIÊN (ĐÃ SỬA THEO YÊU CẦU)
    # ════════════════════════════════════════════════════════════

    # modules/tkb/export_service.py - Hàm _create_sheet_tkb_gv hoàn chỉnh

    def _create_sheet_tkb_gv(self, wb, nam_hoc_id, hoc_ky_id, thong_tin, nam_hoc_info):
        ws = wb.create_sheet("TKB_Giáo viên")
        
        # ── LẤY DANH SÁCH GIÁO VIÊN ────────────────────────
        tkb_records = self.session.query(ThoiKhoaBieu).filter(
            ThoiKhoaBieu.nam_hoc_id == nam_hoc_id,
            ThoiKhoaBieu.hoc_ky_id == hoc_ky_id,
            ThoiKhoaBieu.is_active == True
        ).all()
        
        gv_ids = sorted(list(set([t.giao_vien_id for t in tkb_records])))
        gv_list = []
        for gv_id in gv_ids:
            gv = self.session.query(GiaoVien).filter(GiaoVien.id == gv_id).first()
            if gv:
                gv_list.append(gv)
        
        if not gv_list:
            ws['A1'] = "Không có dữ liệu TKB"
            return
        
        # ── HEADER ──────────────────────────────────────────
        row = 1
        total_cols = 2 + len(gv_list) * 2
        last_col = get_column_letter(total_cols)
        
        # Dòng 1: Tên trường
        ws.merge_cells(f'A{row}:{last_col}{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = thong_tin['ten_truong'].upper()
        cell.font = Font(name='Times New Roman', size=14, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # Dòng 2: Năm học - Học kỳ
        ws.merge_cells(f'A{row}:{last_col}{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = f"Năm học {nam_hoc_info['ten_nam_hoc']} - {nam_hoc_info['ten_hoc_ky']}"
        cell.font = Font(name='Times New Roman', size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # Dòng 3: Tiêu đề
        ws.merge_cells(f'A{row}:{last_col}{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = "THỜI KHÓA BIỂU GIÁO VIÊN"
        cell.font = Font(name='Times New Roman', size=14, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 2
        
        # ── HEADER BẢNG ────────────────────────────────────
        # Dòng 1: THỨ | TIẾT | GV1 | GV2 | ...
        headers = ['THỨ', 'TIẾT']
        gv_cols = {}
        
        for idx, gv in enumerate(gv_list):
            ten_gv = gv.nguoi_dung.ho_ten if gv.nguoi_dung else f"GV {gv.id}"
            col_start = 2 + idx * 2 + 1
            col_end = col_start + 1
            gv_cols[gv.id] = (col_start, col_end)
            
            # ⭐ Merge 2 cột cho tên GV
            ws.merge_cells(
                start_row=row, 
                start_column=col_start, 
                end_row=row, 
                end_column=col_end
            )
            cell = ws.cell(row=row, column=col_start)
            cell.value = ten_gv
            cell.font = Font(name='Times New Roman', size=10, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            # ⭐ Viền cho ô merge (kẻ toàn bộ viền)
            self._apply_border_to_cell(ws, row, col_start)
            # ⭐ Kẻ thêm viền cho ô bên phải (col_end) để viền trên cùng được kẻ đầy đủ
            self._apply_border_to_cell(ws, row, col_end)
        
        # Cột THỨ và TIẾT
        cell = ws.cell(row=row, column=1)
        cell.value = 'THỨ'
        cell.font = Font(name='Times New Roman', size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        self._apply_border_to_cell(ws, row, 1)
        
        cell = ws.cell(row=row, column=2)
        cell.value = 'TIẾT'
        cell.font = Font(name='Times New Roman', size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        self._apply_border_to_cell(ws, row, 2)
        
        row += 1
        
        # Dòng 2: Sáng | Chiều cho từng GV
        cell = ws.cell(row=row, column=1)
        cell.value = ''
        cell.fill = PatternFill(start_color=self.COLOR_SUB_HEADER, end_color=self.COLOR_SUB_HEADER, fill_type="solid")
        self._apply_border_to_cell(ws, row, 1)
        
        cell = ws.cell(row=row, column=2)
        cell.value = ''
        cell.fill = PatternFill(start_color=self.COLOR_SUB_HEADER, end_color=self.COLOR_SUB_HEADER, fill_type="solid")
        self._apply_border_to_cell(ws, row, 2)
        
        for idx, gv in enumerate(gv_list):
            col_start, col_end = gv_cols[gv.id]
            
            # Cột Sáng
            cell = ws.cell(row=row, column=col_start)
            cell.value = 'Sáng'
            cell.font = Font(name='Times New Roman', size=9, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color=self.COLOR_SUB_HEADER, end_color=self.COLOR_SUB_HEADER, fill_type="solid")
            self._apply_border_to_cell(ws, row, col_start)
            
            # Cột Chiều
            cell = ws.cell(row=row, column=col_end)
            cell.value = 'Chiều'
            cell.font = Font(name='Times New Roman', size=9, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color=self.COLOR_SUB_HEADER, end_color=self.COLOR_SUB_HEADER, fill_type="solid")
            self._apply_border_to_cell(ws, row, col_end)
        
        row += 1
        
        # ── DỮ LIỆU ────────────────────────────────────────
        tkb_map = {}
        for t in tkb_records:
            key = (t.giao_vien_id, t.thu, t.tiet, t.buoi)
            tkb_map[key] = t
        
        thu_list = [2, 3, 4, 5, 6, 7]
        thu_names = {2: '2', 3: '3', 4: '4', 5: '5', 6: '6', 7: '7'}
        max_tiet = 5
        start_row = row
        
        for thu in thu_list:
            for tiet in range(1, max_tiet + 1):
                # ⭐ Cột THỨ (GỘP Ô CHO CÁC TIẾT CÙNG THỨ)
                if tiet == 1:
                    # Merge cột THỨ cho 5 tiết của cùng thứ
                    ws.merge_cells(
                        start_row=row, 
                        start_column=1, 
                        end_row=row + max_tiet - 1, 
                        end_column=1
                    )
                    cell = ws.cell(row=row, column=1)
                    cell.value = thu_names[thu]
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(name='Times New Roman', size=10, bold=True)
                    # ⭐ Kẻ viền cho ô đã merge
                    self._apply_border_to_cell(ws, row, 1)
                    self._apply_border_to_cell(ws, row + max_tiet - 1, 1)
                
                # Cột TIẾT
                cell = ws.cell(row=row, column=2)
                cell.value = f"{tiet}"
                cell.alignment = Alignment(horizontal='center', vertical='center')
                self._apply_border_to_cell(ws, row, 2)
                
                # Dữ liệu từng giáo viên (Sáng + Chiều)
                for gv in gv_list:
                    col_start, col_end = gv_cols[gv.id]
                    
                    # Buổi Sáng
                    key_sang = (gv.id, thu, tiet, 'sang')
                    tkb_sang = tkb_map.get(key_sang)
                    if tkb_sang:
                        mon = tkb_sang.mon_hoc
                        lop = tkb_sang.lop_hoc
                        value = f"{self._ten_mon(tkb_sang)} - {lop.ten_lop if lop else ''}"
                    else:
                        value = ''
                    cell = ws.cell(row=row, column=col_start)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    self._apply_border_to_cell(ws, row, col_start)
                    
                    # Buổi Chiều
                    key_chieu = (gv.id, thu, tiet, 'chieu')
                    tkb_chieu = tkb_map.get(key_chieu)
                    if tkb_chieu:
                        mon = tkb_chieu.mon_hoc
                        lop = tkb_chieu.lop_hoc
                        value = f"{self._ten_mon(tkb_chieu)} - {lop.ten_lop if lop else ''}"
                    else:
                        value = ''
                    cell = ws.cell(row=row, column=col_end)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    self._apply_border_to_cell(ws, row, col_end)
                
                row += 1
        
        # ⭐ Điều chỉnh chiều cao hàng cho dễ đọc
        for r in range(1, row):
            ws.row_dimensions[r].height = 20
        
        # ── ĐỘ RỘNG CỘT ────────────────────────────────────
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 10
        for i in range(3, total_cols + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20

# ═══════════════════════════════════════════════════════════════
#  THÊM HÀM ÁP DỤNG VIỀN CHO TOÀN BỘ BẢNG
# ═══════════════════════════════════════════════════════════════

    def _apply_border_to_range(self, ws, start_row, start_col, end_row, end_col):
        """Áp dụng viền cho toàn bộ vùng dữ liệu"""
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = border

    def _apply_border_to_cell(self, ws, row, col):
        """Áp dụng viền cho 1 ô"""
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        ws.cell(row=row, column=col).border = border
    def _get_ten_mon_hien_thi(self, phan_cong_id: int = None, mon_hoc=None, lop_hoc=None):
        """
        Lấy tên môn hiển thị:
        - Nếu là môn có phân môn và có phan_cong_id → lấy tên phân môn
        - Nếu có lop_hoc, tìm phân công của môn đó cho lớp đó
        - Ngược lại → lấy tên môn chính
        """
        if not mon_hoc:
            return ''
        
        # Nếu môn không có phân môn → trả về tên môn
        if not mon_hoc.co_phan_mon:
            return mon_hoc.ten_mon
        
        # Cách 1: Nếu có phan_cong_id
        if phan_cong_id:
            try:
                from core.db.models.phan_cong import PhanCongGiangDay
                from core.db.models.mon_hoc import PhanMon
                pc = self.session.query(PhanCongGiangDay).filter(
                    PhanCongGiangDay.id == phan_cong_id
                ).first()
                if pc and pc.phan_mon_id:
                    pm = self.session.query(PhanMon).filter(
                        PhanMon.id == pc.phan_mon_id
                    ).first()
                    if pm:
                        return pm.ten_phan_mon
            except Exception as e:
                print(f"⚠️ Lỗi lấy phân môn từ phan_cong_id: {e}")
        
        # Cách 2: Nếu có lop_hoc, tìm phân công của môn đó cho lớp đó
        if lop_hoc:
            try:
                from core.db.models.phan_cong import PhanCongGiangDay
                from core.db.models.mon_hoc import PhanMon
                pc = self.session.query(PhanCongGiangDay).filter(
                    PhanCongGiangDay.mon_hoc_id == mon_hoc.id,
                    PhanCongGiangDay.lop_hoc_id == lop_hoc.id
                ).first()
                if pc and pc.phan_mon_id:
                    pm = self.session.query(PhanMon).filter(
                        PhanMon.id == pc.phan_mon_id
                    ).first()
                    if pm:
                        return pm.ten_phan_mon
            except Exception as e:
                print(f"⚠️ Lỗi lấy phân môn từ lop_hoc: {e}")
        
        # Mặc định trả về tên môn
        return mon_hoc.ten_mon

    def _ten_mon(self, tkb):
        """Lấy tên môn từ bản ghi TKB"""
        if not tkb:
            return ''
        mon = tkb.mon_hoc
        if not mon:
            return ''
        # Sử dụng hàm _get_ten_mon_hien_thi
        return self._get_ten_mon_hien_thi(
            phan_cong_id=getattr(tkb, 'phan_cong_id', None),
            mon_hoc=mon,
            lop_hoc=tkb.lop_hoc
        )