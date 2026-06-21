from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
                               QTableWidget, QTableWidgetItem, QMessageBox,
                               QAbstractItemView, QHeaderView, QLabel, QMenu)
from PySide6.QtCore import Qt
from core.db.session import SessionLocal
from modules.phan_cong.service import PhanCongService
from modules.phan_cong.dialogs.phan_cong_dialog import PhanCongDialog


class PhanCongPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.session = SessionLocal()
        self.svc = PhanCongService(self.session)
        self._build_ui()
        self._load()
    
    def closeEvent(self, event):
        self.session.close()
        super().closeEvent(event)
    
    def _build_ui(self):
        layout = QVBoxLayout(self)
        
        # Toolbar
        tb = QHBoxLayout()
        self.btn_them = QPushButton("+ Phân công")
        self.btn_export = QPushButton("📊 Xuất Excel")
        self.btn_export.setObjectName("btn_success")
        self.btn_export.clicked.connect(self._export_excel)
        self.btn_them.clicked.connect(self._them)
        
        # Nút Xóa có menu thả xuống
        self.btn_xoa = QPushButton("🗑 Xóa")
        self.btn_xoa.setEnabled(False)
        
        # Tạo menu cho nút Xóa
        self.xoa_menu = QMenu(self)
        self.xoa_chon = self.xoa_menu.addAction("🗑 Xóa phân công đang chọn")
        self.xoa_tat_ca = self.xoa_menu.addAction("⚠️ Xóa TẤT CẢ phân công")
        
        # Kết nối menu
        self.xoa_chon.triggered.connect(self._xoa_chon)
        self.xoa_tat_ca.triggered.connect(self._xoa_tat_ca)
        
        # Gán menu cho nút Xóa
        self.btn_xoa.setMenu(self.xoa_menu)
        # Click chuột trái vẫn xóa cái đang chọn
        self.btn_xoa.clicked.connect(self._xoa_chon)
        
        tb.addWidget(self.btn_them)
        tb.addWidget(self.btn_xoa)
        tb.addWidget(self.btn_export)
        tb.addStretch()
        layout.addLayout(tb)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["ID", "Năm học", "Học kỳ", "Giáo viên", "Môn học", "Lớp học"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.table.itemSelectionChanged.connect(self._on_selection)
        layout.addWidget(self.table)
    
    def _load(self):
        r = self.svc.lay_ds()
        if r.ok:
            self._render(r.data)
    
    def _render(self, data):
        # Hàm lấy số thứ tự khối để sắp xếp đúng
        def get_khoi_order(lop):
            if lop:
                return lop.khoi
            return 99
        
        # Hàm lấy tên lớp để sắp xếp (6A, 6B, 7A...)
        def get_lop_name(lop):
            if lop:
                return lop.ten_lop
            return ""
        
        sorted_data = sorted(data, key=lambda x: (
            x.giao_vien.nguoi_dung.ho_ten if x.giao_vien and x.giao_vien.nguoi_dung else "",
            x.mon_hoc.ten_mon if x.mon_hoc else "",
            get_khoi_order(x.lop_hoc),
            get_lop_name(x.lop_hoc)
        ))
        
        self.table.setRowCount(0)
        for row, item in enumerate(sorted_data):
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(str(item.id)))
            self.table.setItem(row, 1, QTableWidgetItem(item.nam_hoc.ten_nam_hoc if item.nam_hoc else ""))
            self.table.setItem(row, 2, QTableWidgetItem(item.hoc_ky.ten_hoc_ky if item.hoc_ky else ""))
            self.table.setItem(row, 3, QTableWidgetItem(item.giao_vien.nguoi_dung.ho_ten if item.giao_vien and item.giao_vien.nguoi_dung else ""))
            self.table.setItem(row, 4, QTableWidgetItem(item.mon_hoc.ten_mon if item.mon_hoc else ""))
            self.table.setItem(row, 5, QTableWidgetItem(item.lop_hoc.ten_lop if item.lop_hoc else ""))
            self.table.item(row, 0).setData(Qt.UserRole, item.id)
    
    def _on_selection(self):
        has = bool(self.table.selectedItems())
        self.btn_xoa.setEnabled(has)
    
    def _get_selected_id(self):
        row = self.table.currentRow()
        if row >= 0:
            return self.table.item(row, 0).data(Qt.UserRole)
        return None
    
    def _them(self):
        ds_nam_hoc = self.svc.lay_ds_nam_hoc()
        ds_hoc_ky = self.svc.lay_ds_hoc_ky()
        ds_giao_vien = self.svc.lay_ds_giao_vien()
        ds_mon_hoc = self.svc.lay_ds_mon_hoc()
        ds_lop_hoc = self.svc.lay_ds_lop_hoc()
        
        # Hàm lấy phân công hiện tại theo GV, NH, HK
        def lay_phan_cong_theo_gv(gv_id, nam_hoc_id, hoc_ky_id):
            return self.svc.lay_phan_cong_theo_giao_vien(gv_id, nam_hoc_id, hoc_ky_id)
        
        dlg = PhanCongDialog(
            self, 
            ds_nam_hoc=ds_nam_hoc, 
            ds_hoc_ky=ds_hoc_ky,
            ds_giao_vien=ds_giao_vien, 
            ds_mon_hoc=ds_mon_hoc,
            ds_lop_hoc=ds_lop_hoc,
            lay_phan_cong_theo_gv_func=lay_phan_cong_theo_gv
        )
        
        if dlg.exec() == PhanCongDialog.Accepted:
            r = self.svc.them(**dlg.get_data())
            if r.ok:
                self._load()
                QMessageBox.information(self, "Thành công", r.error)
            else:
                QMessageBox.warning(self, "Lỗi", r.error)
    
    def _xoa_chon(self):
        """Xóa phân công đang chọn"""
        id = self._get_selected_id()
        if not id:
            QMessageBox.warning(self, "Lỗi", "Vui lòng chọn phân công cần xóa.")
            return
        
        reply = QMessageBox.question(self, "Xác nhận", "Xóa phân công này?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            r = self.svc.xoa(id)
            if r.ok:
                self._load()
                QMessageBox.information(self, "Thành công", r.error)
            else:
                QMessageBox.warning(self, "Lỗi", r.error)
    
    def _xoa_tat_ca(self):
        """Xóa tất cả phân công trong bảng hiện tại"""
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "Lỗi", "Không có phân công nào để xóa.")
            return
        
        reply = QMessageBox.question(
            self, 
            "Xác nhận xóa tất cả",
            f"Bạn có chắc chắn muốn xóa TOÀN BỘ {self.table.rowCount()} phân công?\n"
            "Hành động này không thể hoàn tác.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            # Xóa từng cái một hoặc xóa tất cả bằng service
            # Cách 1: Xóa từng cái
            success_count = 0
            fail_count = 0
            for row in range(self.table.rowCount()):
                item = self.table.item(row, 0)
                if item:
                    id = item.data(Qt.UserRole)
                    r = self.svc.xoa(id)
                    if r.ok:
                        success_count += 1
                    else:
                        fail_count += 1
            
            self._load()
            QMessageBox.information(
                self, 
                "Kết quả xóa", 
                f"Đã xóa thành công {success_count} phân công.\n"
                f"Thất bại: {fail_count}"
            )
    
    def _export_excel(self):
        """Xuất danh sách phân công ra Excel và tự động mở"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from datetime import datetime
            import os
            import subprocess
            
            # Lấy dữ liệu
            r = self.svc.lay_ds_phan_cong_tong_hop()
            if not r.ok:
                QMessageBox.warning(self, "Lỗi", r.error)
                return
            
            data = r.data
            
            # Tạo workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Phân công giảng dạy"
            
            # Định dạng
            header_fill = PatternFill(start_color="1D9E75", end_color="1D9E75", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Tiêu đề
            ws.merge_cells('A1:D1')
            title = ws['A1']
            title.value = f"BẢNG PHÂN CÔNG GIẢNG DẠY - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            title.font = Font(size=14, bold=True)
            title.alignment = Alignment(horizontal='center')
            
            # Header
            headers = ["STT", "Họ và tên", "Phân công giảng dạy", "Tổng số tiết"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Dữ liệu
            for row, item in enumerate(data, 1):
                # STT
                ws.cell(row=row+3, column=1, value=row).border = thin_border
                ws.cell(row=row+3, column=1).alignment = Alignment(horizontal='center')
                
                # Họ tên
                ws.cell(row=row+3, column=2, value=item['ho_ten']).border = thin_border
                
                # Phân công giảng dạy
                phan_cong_str = " + ".join([
                    f"{m['mon_hoc']} ({','.join(m['lops'])})" 
                    for m in item['phan_cong']
                ])
                ws.cell(row=row+3, column=3, value=phan_cong_str).border = thin_border
                
                # Tổng số tiết
                ws.cell(row=row+3, column=4, value=item['tong_tiet']).border = thin_border
                ws.cell(row=row+3, column=4).alignment = Alignment(horizontal='center')
            
            # Điều chỉnh độ rộng cột
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 15
            
            # Tạo thư mục nếu chưa có
            export_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "data", "exports")
            os.makedirs(export_dir, exist_ok=True)
            
            # Lưu file
            file_name = f"phan_cong_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(export_dir, file_name)
            
            wb.save(file_path)
            
            # Tự động mở file
            os.startfile(file_path)
            
            QMessageBox.information(self, "Thành công", f"Đã xuất và mở file:\n{file_path}")
            
        except ImportError:
            QMessageBox.warning(self, "Lỗi", "Chưa cài đặt openpyxl. Chạy: pip install openpyxl")
        except Exception as e:
            QMessageBox.warning(self, "Lỗi", f"Xuất file thất bại:\n{str(e)}")