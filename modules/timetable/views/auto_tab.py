# D:\QUANLYTRUONGHOC\modules\timetable\views\auto_tab.py
"""
AutoTimetableTab: tab TKB tự động.
Tính năng:
  - Hiển thị mã môn + tên GV (chỉ tên)
  - Kéo thả giữa các ô
  - Ràng buộc: GV trùng tiết, tiết trống GV, max tiết/ngày, GDTC tiết 5
  - Xuất Excel
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QComboBox, QLabel, QMessageBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView,
    QDialog, QFormLayout, QDialogButtonBox, QSpinBox,
    QCheckBox, QScrollArea, QFrame, QGroupBox
)
from PySide6.QtGui import QColor, QDrag, QFont
from PySide6.QtCore import Qt, QMimeData, QTimer
from core.db.session import SessionLocal
from modules.timetable.service import TimetableService


# ── Hằng số ───────────────────────────────────────────────────
DAYS       = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7"]
THU_NUMS   = [2, 3, 4, 5, 6, 7]
BUOIS      = [("Sáng", True), ("Chiều", False)]
SO_TIET    = 5

COLOR_THU   = QColor("#E8F4FD")
COLOR_SANG  = QColor("#F0FBF6")
COLOR_CHIEU = QColor("#FEF9EC")
COLOR_TIET  = QColor("#F8F9FA")
COLOR_LOCK  = QColor("#90EE90")
COLOR_ERROR = QColor("#FFCCCC")   # vi phạm ràng buộc


# ── Helper: tên tắt ───────────────────────────────────────────
def ten_tac_gv(ho_ten: str) -> str:
    """Lấy tên cuối: 'Nguyễn Văn Hưng' → 'Hưng'"""
    parts = ho_ten.strip().split()
    return parts[-1] if parts else ho_ten


# ══════════════════════════════════════════════════════════════
#  DraggableTable — QTableWidget hỗ trợ kéo thả
# ══════════════════════════════════════════════════════════════
class DraggableTable(QTableWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setDragEnabled(True)
        self.setAcceptDrops(True)
        self.setDropIndicatorShown(True)
        self.setDragDropMode(QAbstractItemView.DragDrop)
        self._drag_source = None   # (row, col) nguồn kéo

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            item = self.itemAt(event.pos())
            if item and item.text().strip():
                self._drag_source = (item.row(), item.column())
        super().mousePressEvent(event)

    def dragEnterEvent(self, event):
        if event.mimeData().hasText():
            event.acceptProposedAction()

    def dragMoveEvent(self, event):
        if event.mimeData().hasText():
            event.acceptProposedAction()

    def startDrag(self, actions):
        if not self._drag_source:
            return
        row, col = self._drag_source
        item = self.item(row, col)
        if not item or not item.text().strip():
            return
        mime = QMimeData()
        mime.setText(f"{row},{col}")
        drag = QDrag(self)
        drag.setMimeData(mime)
        drag.exec(Qt.MoveAction)

    def dropEvent(self, event):
        if not event.mimeData().hasText():
            return
        try:
            src_row, src_col = map(int,
                event.mimeData().text().split(","))
        except ValueError:
            return

        dest = self.itemAt(event.pos())
        if not dest:
            return
        dst_row, dst_col = dest.row(), dest.column()

        # Không drop vào cột header (0,1,2)
        if dst_col < 3 or src_col < 3:
            return
        if src_row == dst_row and src_col == dst_col:
            return

        # Emit signal để parent xử lý swap
        if hasattr(self.parent(), '_on_swap_cells'):
            self.parent()._on_swap_cells(src_row, src_col,
                                          dst_row, dst_col)
        event.acceptProposedAction()


# ══════════════════════════════════════════════════════════════
#  AutoTimetableTab
# ══════════════════════════════════════════════════════════════
class AutoTimetableTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.session        = SessionLocal()
        self.svc            = TimetableService(self.session)
        self._current_data  = {}    # {(thu,buoi,tiet,lop_id): (mon_id,gv_id)}
        self._locked_cells  = set()
        self._ds_lop        = []
        self._ds_mon        = []
        self._ds_gv         = []
        self._lich_hoc      = {}    # {thu: (co_sang, co_chieu)}
        self._rang_buoc     = {     # cấu hình ràng buộc mặc định
            "gv_trung_tiet":    True,
            "max_tiet_ngay":    5,
            "gdtc_khong_tiet5": True,
            "tiet_trong_gv":    {},  # {gv_id: [(thu, buoi, tiet)]}
        }
        self._row_map       = []    # list row_data theo thứ tự render
        self._build_ui()
        self._load_filters()
        QTimer.singleShot(100, self._load_existing_timetable)

    # ── Build UI ──────────────────────────────────────────────
    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(8)
        layout.setContentsMargins(8, 8, 8, 8)

        # === HÀNG 1: Bộ lọc ===
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(8)
        
        filter_layout.addWidget(QLabel("Năm học:"))
        self.cmb_nam_hoc = QComboBox()
        self.cmb_nam_hoc.setFixedWidth(120)
        self.cmb_nam_hoc.currentIndexChanged.connect(self._on_nam_hoc_changed)
        filter_layout.addWidget(self.cmb_nam_hoc)
        
        filter_layout.addWidget(QLabel("Học kỳ:"))
        self.cmb_hoc_ky = QComboBox()
        self.cmb_hoc_ky.setFixedWidth(100)
        filter_layout.addWidget(self.cmb_hoc_ky)
        
        filter_layout.addWidget(QLabel("Lớp:"))
        self.cmb_filter_lop = QComboBox()
        self.cmb_filter_lop.setFixedWidth(80)
        self.cmb_filter_lop.addItem("Tất cả", None)
        self.cmb_filter_lop.currentIndexChanged.connect(self._render_table)
        filter_layout.addWidget(self.cmb_filter_lop)
        
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # === HÀNG 2: Nút chức năng (hàng trên) ===
        row1 = QHBoxLayout()
        row1.setSpacing(6)
        
        for text, slot in [
            ("📅 Lịch học", self._cau_hinh_lich_hoc),
            ("⚙️ Ràng buộc", self._cau_hinh_rang_buoc),
            ("🚀 Sinh TKB", self._generate),
        ]:
            btn = QPushButton(text)
            btn.setFixedHeight(30)
            btn.setMinimumWidth(90)
            btn.setStyleSheet("""
                QPushButton {
                    background: #E8F0FE;
                    border-radius: 5px;
                    padding: 4px 12px;
                    font-weight: 500;
                }
                QPushButton:hover {
                    background: #D0E0F0;
                }
            """)
            btn.clicked.connect(slot)
            row1.addWidget(btn)
        
        row1.addStretch()
        layout.addLayout(row1)

        # === HÀNG 3: Nút chức năng (hàng dưới) ===
        row2 = QHBoxLayout()
        row2.setSpacing(6)
        
        for text, slot, is_primary in [
            ("🔒 Khóa", self._lock_selected, False),
            ("🔓 Mở khóa", self._unlock_selected, False),
            ("✅ Kiểm tra", self._kiem_tra_rang_buoc, False),
            ("🔍 Tổng thể", self._kiem_tra_tong_the, False),
            ("💾 Lưu", self._save_all, False),
            ("📊 Xuất Excel", self._xuat_excel, True),
        ]:
            btn = QPushButton(text)
            btn.setFixedHeight(30)
            btn.setMinimumWidth(90)
            if is_primary:
                btn.setStyleSheet("""
                    QPushButton {
                        background: #217346;
                        color: white;
                        border-radius: 5px;
                        padding: 4px 12px;
                        font-weight: 500;
                    }
                    QPushButton:hover {
                        background: #1B5E3A;
                    }
                """)
            else:
                btn.setStyleSheet("""
                    QPushButton {
                        background: #F5F5F5;
                        border-radius: 5px;
                        padding: 4px 12px;
                    }
                    QPushButton:hover {
                        background: #E0E0E0;
                    }
                """)
            btn.clicked.connect(slot)
            row2.addWidget(btn)
        
        row2.addStretch()
        layout.addLayout(row2)

        # === Bảng TKB ===
        self.table = DraggableTable(self)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.itemDoubleClicked.connect(self._on_cell_double_click)
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.table.setMinimumHeight(500)
        layout.addWidget(self.table)
        
        # === Thanh trạng thái ===
        status_layout = QHBoxLayout()
        self.lbl_status = QLabel("✅ Sẵn sàng")
        self.lbl_status.setStyleSheet("color: #888; font-size: 11px;")
        status_layout.addWidget(self.lbl_status)
        status_layout.addStretch()
        
        note = QLabel("💡 Double-click: sửa | Kéo thả: hoán đổi | Xanh: khóa | Đỏ: lỗi")
        note.setStyleSheet("color: #aaa; font-size: 10px;")
        status_layout.addWidget(note)
        
        layout.addLayout(status_layout)
    # ── Load dữ liệu ─────────────────────────────────────────
    def _load_filters(self):
        self._ds_lop = self.svc.lay_ds_lop()
        self._ds_mon = self.svc.lay_ds_mon()
        self._ds_gv  = self.svc.lay_ds_giao_vien()

        from core.db.models.nam_hoc import NamHoc
        for nh in self.session.query(NamHoc).filter_by(active=True).all():
            self.cmb_nam_hoc.addItem(nh.ten_nam_hoc, nh.id)

        # Filter lớp
        self.cmb_filter_lop.clear()
        self.cmb_filter_lop.addItem("Tất cả", None)
        for lop in self._ds_lop:
            self.cmb_filter_lop.addItem(lop.ten_lop, lop.id)

    def _on_nam_hoc_changed(self):
        self.cmb_hoc_ky.clear()
        nam_hoc_id = self.cmb_nam_hoc.currentData()
        if nam_hoc_id:
            from core.db.models.hoc_ky import HocKy
            for hk in self.session.query(HocKy).filter_by(
                    nam_hoc_id=nam_hoc_id, active=True).all():
                self.cmb_hoc_ky.addItem(hk.ten_hoc_ky, hk.id)
        if self.cmb_hoc_ky.count() > 0:
            self._load_lich_hoc()
            self._load_existing_timetable()

    # ── Lịch học ─────────────────────────────────────────────
    def _load_lich_hoc(self):
        try:
            from core.db.models.lich_hoc_tuan import LichHocTuan
            nam_hoc_id = self.cmb_nam_hoc.currentData()
            hoc_ky_id  = self.cmb_hoc_ky.currentData()
            if not nam_hoc_id or not hoc_ky_id:
                self._lich_hoc = {t: (True, False) for t in THU_NUMS}
                return
            ds = self.session.query(LichHocTuan).filter_by(
                nam_hoc_id=nam_hoc_id, hoc_ky_id=hoc_ky_id).all()
            self._lich_hoc = ({lht.thu: (lht.co_sang, lht.co_chieu)
                               for lht in ds}
                              if ds else {t: (True, False) for t in THU_NUMS})
        except Exception:
            self._lich_hoc = {t: (True, False) for t in THU_NUMS}

    def _cau_hinh_lich_hoc(self):
        try:
            from modules.timetable.dialogs.lich_hoc_tuan_dialog import (
                LichHocTuanDialog)
            from core.db.models.lich_hoc_tuan import LichHocTuan
        except ImportError:
            QMessageBox.warning(self, "Lỗi",
                                "Chưa có file lich_hoc_tuan_dialog.py")
            return

        nam_hoc_id = self.cmb_nam_hoc.currentData()
        hoc_ky_id  = self.cmb_hoc_ky.currentData()
        if not nam_hoc_id or not hoc_ky_id:
            QMessageBox.warning(self, "Thiếu",
                                "Chọn năm học và học kỳ trước.")
            return
        lich_db = self.session.query(LichHocTuan).filter_by(
            nam_hoc_id=nam_hoc_id, hoc_ky_id=hoc_ky_id).all()
        dlg = LichHocTuanDialog(self, nam_hoc_id=nam_hoc_id,
                                hoc_ky_id=hoc_ky_id,
                                lich_hien_tai=lich_db)
        if dlg.exec() != QDialog.Accepted:
            return
        self.session.query(LichHocTuan).filter_by(
            nam_hoc_id=nam_hoc_id, hoc_ky_id=hoc_ky_id).delete()
        for row in dlg.get_data():
            self.session.add(LichHocTuan(**row))
        self.session.commit()
        self._lich_hoc = dlg.get_lich_dict()
        self._render_table()

    # ── Ràng buộc ─────────────────────────────────────────────
    def _cau_hinh_rang_buoc(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Cấu hình ràng buộc TKB")
        dlg.setMinimumWidth(400)
        layout = QVBoxLayout(dlg)
        layout.setSpacing(10)

        # GV trùng tiết
        chk_gv = QCheckBox(
            "GV không dạy 2 lớp cùng 1 tiết")
        chk_gv.setChecked(self._rang_buoc["gv_trung_tiet"])
        layout.addWidget(chk_gv)

        # Max tiết/ngày
        row_max = QHBoxLayout()
        chk_max = QCheckBox("Không xếp quá")
        chk_max.setChecked(self._rang_buoc["max_tiet_ngay"] > 0)
        spn_max = QSpinBox()
        spn_max.setRange(1, 10)
        spn_max.setValue(self._rang_buoc["max_tiet_ngay"] or 5)
        row_max.addWidget(chk_max)
        row_max.addWidget(spn_max)
        row_max.addWidget(QLabel("tiết/ngày cho 1 lớp"))
        row_max.addStretch()
        layout.addLayout(row_max)

        # GDTC không tiết 5
        chk_gdtc = QCheckBox(
            "GDTC/Thể dục không xếp tiết 5 buổi sáng")
        chk_gdtc.setChecked(self._rang_buoc["gdtc_khong_tiet5"])
        layout.addWidget(chk_gdtc)

        # Tiết trống GV (đơn giản: nhập dạng text)
        grp = QGroupBox("Tiết trống bắt buộc của GV")
        grp_layout = QVBoxLayout(grp)
        note = QLabel(
            "Nhập mỗi dòng: MãGV Thứ Buổi Tiết\n"
            "VD: GV001 3 Sáng 1\n"
            "    GV001 3 Sáng 2  (GV001 nghỉ sáng thứ 3 tiết 1,2)")
        note.setStyleSheet("color:#7A8BAD; font-size:11px;")
        grp_layout.addWidget(note)

        from PySide6.QtWidgets import QTextEdit
        txt_tiet_trong = QTextEdit()
        txt_tiet_trong.setMaximumHeight(100)
        txt_tiet_trong.setPlaceholderText("GV001 3 Sáng 1\nGV002 5 Chiều 3")
        # Fill dữ liệu hiện tại
        lines = []
        for gv_id, slots in self._rang_buoc["tiet_trong_gv"].items():
            gv = self._get_gv_by_id(gv_id)
            ma = gv.ma_giao_vien if gv else str(gv_id)
            for thu, buoi, tiet in slots:
                lines.append(f"{ma} {thu} {buoi} {tiet}")
        txt_tiet_trong.setPlainText("\n".join(lines))
        grp_layout.addWidget(txt_tiet_trong)
        layout.addWidget(grp)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.button(QDialogButtonBox.Ok).setText("Lưu")
        btns.button(QDialogButtonBox.Cancel).setText("Huỷ")
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        layout.addWidget(btns)

        if dlg.exec() != QDialog.Accepted:
            return

        self._rang_buoc["gv_trung_tiet"]    = chk_gv.isChecked()
        self._rang_buoc["max_tiet_ngay"]    = (spn_max.value()
                                               if chk_max.isChecked() else 0)
        self._rang_buoc["gdtc_khong_tiet5"] = chk_gdtc.isChecked()

        # Parse tiết trống GV
        tiet_trong = {}
        for line in txt_tiet_trong.toPlainText().strip().split("\n"):
            parts = line.strip().split()
            if len(parts) < 4:
                continue
            ma_gv, thu_s, buoi, tiet_s = parts[0], parts[1], parts[2], parts[3]
            gv = next((g for g in self._ds_gv
                       if g.ma_giao_vien == ma_gv), None)
            if gv:
                tiet_trong.setdefault(gv.id, [])
                try:
                    tiet_trong[gv.id].append(
                        (int(thu_s), buoi, int(tiet_s)))
                except ValueError:
                    pass
        self._rang_buoc["tiet_trong_gv"] = tiet_trong
        QMessageBox.information(self, "Đã lưu", "Cấu hình ràng buộc đã lưu.")

    def _vi_pham_rang_buoc(self, key, mon_id, gv_id) -> list[str]:
        """Kiểm tra 1 ô có vi phạm ràng buộc không. Trả về list lỗi."""
        thu, buoi, tiet, lop_id = key
        loi = []

        # 1. GV trùng tiết
        if self._rang_buoc["gv_trung_tiet"]:
            for (t, b, ti, l), (m, g) in self._current_data.items():
                if (t == thu and b == buoi and ti == tiet
                        and l != lop_id and g == gv_id):
                    lop = self._get_lop_name(l)
                    loi.append(f"GV đang dạy lớp {lop} cùng tiết!")

        # 2. Max tiết/ngày
        max_t = self._rang_buoc["max_tiet_ngay"]
        if max_t > 0:
            count = sum(1 for (t, b, ti, l), _ in self._current_data.items()
                        if t == thu and l == lop_id)
            if count >= max_t:
                loi.append(f"Lớp đã có {count} tiết ngày này (tối đa {max_t})!")

        # 3. GDTC không tiết 5
        if self._rang_buoc["gdtc_khong_tiet5"] and tiet == 5 and buoi == "Sáng":
            mon = self._get_mon_by_id(mon_id)
            if mon and any(k in mon.ma_mon.upper()
                           for k in ["GDTC", "TD", "THE DUC"]):
                loi.append("GDTC không được xếp tiết 5 sáng!")

        # 4. Tiết trống GV
        tiet_trong = self._rang_buoc["tiet_trong_gv"].get(gv_id, [])
        if (thu, buoi, tiet) in [(t, b, ti) for t, b, ti in tiet_trong]:
            loi.append("GV đã đăng ký tiết trống bắt buộc!")

        return loi
         # 5. KIỂM TRA SỐ TIẾT PHÂN BỐ ĐÚNG VỚI YÊU CẦU
        # Lấy số tiết yêu cầu của môn học theo khối
        lop = next((l for l in self._ds_lop if l.id == lop_id), None)
        if lop:
            from core.db.models.mon_hoc_khoi import MonHocKhoi
            so_tiet_yeu_cau = self.session.query(MonHocKhoi).filter(
                MonHocKhoi.mon_hoc_id == mon_id,
                MonHocKhoi.khoi == lop.khoi
            ).first()
            tiet_can = so_tiet_yeu_cau.so_tiet if so_tiet_yeu_cau else 0
            
            # Đếm số tiết hiện tại của môn này trong lớp này
            tiet_da_phan = sum(1 for (t, b, ti, l), (m, g) in self._current_data.items()
                            if l == lop_id and m == mon_id)
            
            # Nếu đã vượt quá số tiết yêu cầu
            if tiet_da_phan > tiet_can:
                loi.append(f"Môn chỉ được {tiet_can} tiết/tuần, đã phân {tiet_da_phan} tiết!")

        return loi
    def _kiem_tra_rang_buoc(self):
        """Highlight đỏ tất cả ô vi phạm."""
        vi_pham_count = 0
        for row_idx, row_data in enumerate(self._row_map):
            thu  = row_data["thu"]
            buoi = row_data["buoi"]
            tiet = row_data["tiet"]
            for col_i, lop in enumerate(self._visible_lop()):
                key = (thu, buoi, tiet, lop.id)
                cell = self.table.item(row_idx, 3 + col_i)
                if not cell:
                    continue
                if key in self._current_data:
                    mon_id, gv_id = self._current_data[key]
                    loi = self._vi_pham_rang_buoc(key, mon_id, gv_id)
                    if loi:
                        cell.setBackground(COLOR_ERROR)
                        cell.setToolTip("\n".join(loi))
                        vi_pham_count += 1
                    elif key in self._locked_cells:
                        cell.setBackground(COLOR_LOCK)
                    else:
                        cell.setBackground(Qt.white)

        if vi_pham_count == 0:
            QMessageBox.information(self, "✅ Không vi phạm",
                                    "TKB không có ràng buộc nào bị vi phạm!")
        else:
            QMessageBox.warning(self, f"⚠️ {vi_pham_count} vi phạm",
                                f"Có {vi_pham_count} ô vi phạm ràng buộc "
                                f"(highlight đỏ). Hover để xem chi tiết.")

    def _kiem_tra_tong_the(self):
        """Kiểm tra toàn bộ TKB có đúng số tiết không"""
        errors = []
        for lop in self._ds_lop:
            for mon in self._ds_mon:
                from core.db.models.mon_hoc_khoi import MonHocKhoi
                so_tiet_yeu_cau = self.session.query(MonHocKhoi).filter(
                    MonHocKhoi.mon_hoc_id == mon.id,
                    MonHocKhoi.khoi == lop.khoi
                ).first()
                tiet_can = so_tiet_yeu_cau.so_tiet if so_tiet_yeu_cau else 0
                
                tiet_da_phan = sum(1 for (t, b, ti, l), (m, g) in self._current_data.items()
                                if l == lop.id and m == mon.id)
                
                if tiet_da_phan > tiet_can:
                    errors.append(f"⚠️ Lớp {lop.ten_lop} - {mon.ten_mon}: {tiet_da_phan}/{tiet_can} (dư {tiet_da_phan - tiet_can})")
                elif tiet_da_phan < tiet_can:
                    errors.append(f"⚠️ Lớp {lop.ten_lop} - {mon.ten_mon}: {tiet_da_phan}/{tiet_can} (thiếu {tiet_can - tiet_da_phan})")
        
        if errors:
            QMessageBox.warning(self, "Kết quả kiểm tra", "\n".join(errors))
        else:
            QMessageBox.information(self, "Kết quả kiểm tra", "✅ Tất cả môn học đã được phân đúng số tiết!")
    # ── Render bảng ───────────────────────────────────────────
    def _visible_lop(self):
        filter_id = self.cmb_filter_lop.currentData()
        if filter_id:
            return [l for l in self._ds_lop if l.id == filter_id]
        return self._ds_lop

    def _render_table(self):
        if not self._ds_lop:
            return
        if not self._lich_hoc:
            self._load_lich_hoc()

        visible_lop = self._visible_lop()

        # Xây danh sách hàng
        self._row_map = []
        for thu, day in zip(THU_NUMS, DAYS):
            co_sang, co_chieu = self._lich_hoc.get(thu, (True, False))
            for buoi, is_sang in BUOIS:
                if is_sang and not co_sang:
                    continue
                if not is_sang and not co_chieu:
                    continue
                for tiet in range(1, SO_TIET + 1):
                    self._row_map.append({
                        "thu": thu, "day": day,
                        "buoi": buoi, "tiet": tiet,
                        "is_sang": is_sang,
                    })

        n_col = 3 + len(visible_lop)
        self.table.clear()
        self.table.setRowCount(len(self._row_map))
        self.table.setColumnCount(n_col)
        self.table.verticalHeader().setVisible(False)

        headers = ["Thứ/Lớp", "Buổi", "Tiết"] + [
            l.ten_lop for l in visible_lop]
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setColumnWidth(0, 65)
        self.table.setColumnWidth(1, 55)
        self.table.setColumnWidth(2, 35)
        for i in range(len(visible_lop)):
            self.table.horizontalHeader().setSectionResizeMode(
                3 + i, QHeaderView.Stretch)
        for c in range(3):
            self.table.horizontalHeader().setSectionResizeMode(
                c, QHeaderView.Fixed)

        def mc(text, bg=None, bold=False):
            item = QTableWidgetItem(text)
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            if bg:
                item.setBackground(bg)
            if bold:
                f = item.font(); f.setBold(True); item.setFont(f)
            return item

        for row_idx, rd in enumerate(self._row_map):
            thu  = rd["thu"]
            day  = rd["day"]
            buoi = rd["buoi"]
            tiet = rd["tiet"]
            clr  = COLOR_SANG if rd["is_sang"] else COLOR_CHIEU
            self.table.setRowHeight(row_idx, 46)

            # Cột Thứ (merge)
            first_of_thu = next(
                i for i, r in enumerate(self._row_map) if r["thu"] == thu)
            if row_idx == first_of_thu:
                span = sum(1 for r in self._row_map if r["thu"] == thu)
                self.table.setItem(row_idx, 0, mc(day, COLOR_THU, True))
                self.table.setSpan(row_idx, 0, span, 1)

            # Cột Buổi (merge)
            first_of_buoi = next(
                i for i, r in enumerate(self._row_map)
                if r["thu"] == thu and r["buoi"] == buoi)
            if row_idx == first_of_buoi:
                span_b = sum(1 for r in self._row_map
                             if r["thu"] == thu and r["buoi"] == buoi)
                self.table.setItem(row_idx, 1, mc(buoi, clr, True))
                self.table.setSpan(row_idx, 1, span_b, 1)

            # Cột Tiết
            self.table.setItem(row_idx, 2, mc(str(tiet), COLOR_TIET))

            # Cột Lớp
            for col_i, lop in enumerate(visible_lop):
                key = (thu, buoi, tiet, lop.id)
                if key in self._current_data:
                    mon_id, gv_id = self._current_data[key]
                    mon = self._get_mon_by_id(mon_id)
                    gv  = self._get_gv_by_id(gv_id)
                    # Mã môn + tên GV (chỉ tên)
                    ma_mon  = mon.ma_mon  if mon else "?"
                    ten_gv  = ten_tac_gv(
                        gv.nguoi_dung.ho_ten
                        if gv and gv.nguoi_dung else "?")
                    cell = QTableWidgetItem(f"{ma_mon}\n{ten_gv}")
                    cell.setTextAlignment(Qt.AlignCenter)
                    cell.setBackground(
                        COLOR_LOCK if key in self._locked_cells
                        else Qt.white)
                else:
                    cell = QTableWidgetItem("")
                    cell.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row_idx, 3 + col_i, cell)

    # ── Kéo thả swap ─────────────────────────────────────────
    def _on_swap_cells(self, sr, sc, dr, dc):
        if sc < 3 or dc < 3:
            return
        visible_lop = self._visible_lop()

        def get_key(row, col):
            if row >= len(self._row_map):
                return None
            rd = self._row_map[row]
            lop_i = col - 3
            if lop_i < 0 or lop_i >= len(visible_lop):
                return None
            return (rd["thu"], rd["buoi"], rd["tiet"], visible_lop[lop_i].id)

        src_key = get_key(sr, sc)
        dst_key = get_key(dr, dc)
        if not src_key or not dst_key:
            return

        # Swap data
        src_data = self._current_data.get(src_key)
        dst_data = self._current_data.get(dst_key)

        if src_data:
            self._current_data[dst_key] = src_data
        elif dst_key in self._current_data:
            del self._current_data[dst_key]

        if dst_data:
            self._current_data[src_key] = dst_data
        elif src_key in self._current_data:
            del self._current_data[src_key]

        self._render_table()

    # ── Double click sửa ─────────────────────────────────────
    def _on_cell_double_click(self, item):
        row = item.row()
        col = item.column()
        if col < 3 or row >= len(self._row_map):
            return
        visible_lop = self._visible_lop()
        lop_i = col - 3
        if lop_i >= len(visible_lop):
            return

        rd  = self._row_map[row]
        key = (rd["thu"], rd["buoi"], rd["tiet"], visible_lop[lop_i].id)
        cur = self._current_data.get(key, (None, None))
        self._show_edit_dialog(row, col, key, cur)

    def _show_edit_dialog(self, row, col, key, current_data):
        dlg = QDialog(self)
        dlg.setWindowTitle("Chỉnh sửa ô TKB")
        dlg.setMinimumWidth(340)
        layout = QVBoxLayout(dlg)
        form = QFormLayout()

        cmb_mon = QComboBox()
        cmb_mon.addItem("-- Môn --", None)
        for m in self._ds_mon:
            cmb_mon.addItem(f"{m.ma_mon} - {m.ten_mon}", m.id)

        cmb_gv = QComboBox()
        cmb_gv.addItem("-- GV --", None)
        for g in self._ds_gv:
            ten = g.nguoi_dung.ho_ten if g.nguoi_dung else "?"
            cmb_gv.addItem(f"{g.ma_giao_vien} - {ten}", g.id)

        mon_id, gv_id = current_data
        if mon_id:
            idx = cmb_mon.findData(mon_id)
            if idx >= 0: cmb_mon.setCurrentIndex(idx)
        if gv_id:
            idx = cmb_gv.findData(gv_id)
            if idx >= 0: cmb_gv.setCurrentIndex(idx)

        form.addRow("Môn:", cmb_mon)
        form.addRow("GV:",  cmb_gv)
        layout.addLayout(form)

        lbl_warn = QLabel()
        lbl_warn.setStyleSheet("color:red; font-size:11px;")
        lbl_warn.setWordWrap(True)
        layout.addWidget(lbl_warn)

        def check():
            m_id = cmb_mon.currentData()
            g_id = cmb_gv.currentData()
            if m_id and g_id:
                loi = self._vi_pham_rang_buoc(key, m_id, g_id)
                lbl_warn.setText("\n".join(loi) if loi else "")

        cmb_mon.currentIndexChanged.connect(check)
        cmb_gv.currentIndexChanged.connect(check)

        btn_box = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
            | QDialogButtonBox.Reset)
        btn_box.button(QDialogButtonBox.Reset).setText("Xóa ô")
        btn_box.accepted.connect(dlg.accept)
        btn_box.rejected.connect(dlg.reject)

        def xoa_o():
            if key in self._current_data:
                del self._current_data[key]
            dlg.accept()
            self._render_table()

        btn_box.button(QDialogButtonBox.Reset).clicked.connect(xoa_o)
        layout.addWidget(btn_box)

        if dlg.exec() == QDialog.Accepted:
            m_id = cmb_mon.currentData()
            g_id = cmb_gv.currentData()
            if m_id and g_id:
                self._current_data[key] = (m_id, g_id)
            elif key in self._current_data:
                del self._current_data[key]
            self._render_table()

    # ── Lock / Unlock ─────────────────────────────────────────
    def _lock_selected(self):
        visible_lop = self._visible_lop()
        for item in self.table.selectedItems():
            row, col = item.row(), item.column()
            if col < 3 or row >= len(self._row_map):
                continue
            rd  = self._row_map[row]
            lop_i = col - 3
            if lop_i < len(visible_lop):
                key = (rd["thu"], rd["buoi"], rd["tiet"],
                       visible_lop[lop_i].id)
                self._locked_cells.add(key)
                item.setBackground(COLOR_LOCK)

    def _unlock_selected(self):
        visible_lop = self._visible_lop()
        for item in self.table.selectedItems():
            row, col = item.row(), item.column()
            if col < 3 or row >= len(self._row_map):
                continue
            rd  = self._row_map[row]
            lop_i = col - 3
            if lop_i < len(visible_lop):
                key = (rd["thu"], rd["buoi"], rd["tiet"],
                       visible_lop[lop_i].id)
                self._locked_cells.discard(key)
                item.setBackground(Qt.white)

    # ── Sinh TKB ─────────────────────────────────────────────
    def _generate(self):
        nam_hoc_id = self.cmb_nam_hoc.currentData()
        hoc_ky_id  = self.cmb_hoc_ky.currentData()
        if not nam_hoc_id or not hoc_ky_id:
            QMessageBox.warning(self, "Lỗi", "Chọn năm học và học kỳ!")
            return

        reply = QMessageBox.question(
            self, "Xác nhận",
            "Sinh TKB mới sẽ xóa dữ liệu cũ (trừ ô khóa). Tiếp tục?",
            QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        from core.db.models.phan_cong import PhanCongGiangDay
        from core.db.models.mon_hoc_khoi import MonHocKhoi
        from core.db.models.lop_hoc import LopHoc
        import random

        # Xoá data cũ (trừ locked)
        self._current_data = {k: v for k, v in self._current_data.items()
                              if k in self._locked_cells}

        if not self._lich_hoc:
            self._load_lich_hoc()

        # Tạo slots trống
        buois_all = [("Sáng", True), ("Chiều", False)]
        for lop in self._ds_lop:
            pc_list = self.session.query(PhanCongGiangDay).filter_by(
                lop_hoc_id=lop.id, nam_hoc_id=nam_hoc_id,
                hoc_ky_id=hoc_ky_id).all()

            mon_list = []
            for pc in pc_list:
                khk = self.session.query(MonHocKhoi).filter_by(
                    mon_hoc_id=pc.mon_hoc_id,
                    khoi=lop.khoi).first()
                n_tiet = khk.so_tiet if khk else 2
                for _ in range(n_tiet):
                    mon_list.append((pc.mon_hoc_id, pc.giao_vien_id))

            slots = []
            for thu in THU_NUMS:
                co_sang, co_chieu = self._lich_hoc.get(thu, (True, False))
                for buoi, is_sang in buois_all:
                    if is_sang and not co_sang:
                        continue
                    if not is_sang and not co_chieu:
                        continue
                    for tiet in range(1, SO_TIET + 1):
                        key = (thu, buoi, tiet, lop.id)
                        if key not in self._locked_cells:
                            slots.append(key)

            random.shuffle(slots)
            random.shuffle(mon_list)

            for i, (mon_id, gv_id) in enumerate(mon_list):
                if i >= len(slots):
                    break
                # Kiểm tra ràng buộc trước khi xếp
                for slot in slots[i:]:
                    loi = self._vi_pham_rang_buoc(slot, mon_id, gv_id)
                    if not loi:
                        self._current_data[slot] = (mon_id, gv_id)
                        slots.remove(slot)
                        break

        self._render_table()
        QMessageBox.information(self, "Thành công",
                                f"Đã sinh TKB cho {len(self._ds_lop)} lớp.")
         # Sau khi sinh xong, kiểm tra lại tổng thể
        # Đếm số tiết đã phân cho từng môn theo lớp
        errors = []
        for lop in self._ds_lop:
            for mon in self._ds_mon:
                so_tiet_yeu_cau = self.session.query(MonHocKhoi).filter(
                    MonHocKhoi.mon_hoc_id == mon.id,
                    MonHocKhoi.khoi == lop.khoi
                ).first()
                tiet_can = so_tiet_yeu_cau.so_tiet if so_tiet_yeu_cau else 0
                
                tiet_da_phan = sum(1 for (t, b, ti, l), (m, g) in self._current_data.items()
                                if l == lop.id and m == mon.id)
                
                if tiet_da_phan != tiet_can:
                    errors.append(f"Lớp {lop.ten_lop} - Môn {mon.ten_mon}: cần {tiet_can} tiết, đã phân {tiet_da_phan} tiết")
        
        if errors:
            msg = "Cảnh báo: Một số môn chưa đúng số tiết!\n- " + "\n- ".join(errors[:10])
            QMessageBox.warning(self, "Cảnh báo", msg)
        else:
            QMessageBox.information(self, "Thành công", f"Đã sinh TKB cho {len(self._ds_lop)} lớp.")
    # ── Lưu DB ────────────────────────────────────────────────
    def _save_all(self):
        nam_hoc_id = self.cmb_nam_hoc.currentData()
        hoc_ky_id  = self.cmb_hoc_ky.currentData()
        if not nam_hoc_id or not hoc_ky_id:
            QMessageBox.warning(self, "Lỗi", "Chọn năm học và học kỳ!")
            return
        from core.db.models.thoi_khoa_bieu import ThoiKhoaBieu
        self.session.query(ThoiKhoaBieu).filter_by(
            nam_hoc_id=nam_hoc_id, hoc_ky_id=hoc_ky_id).delete()
        for (thu, buoi, tiet, lop_id), (mon_id, gv_id) in \
                self._current_data.items():
            tiet_bd = tiet if buoi == "Sáng" else tiet + 5
            self.session.add(ThoiKhoaBieu(
                nam_hoc_id=nam_hoc_id, hoc_ky_id=hoc_ky_id,
                lop_hoc_id=lop_id, mon_hoc_id=mon_id,
                giao_vien_id=gv_id, thu=thu,
                tiet_bat_dau=tiet_bd, so_tiet=1, phong_hoc=""))
        self.session.commit()
        QMessageBox.information(self, "Thành công",
                                f"Đã lưu {len(self._current_data)} tiết.")

    # ── Load từ DB ────────────────────────────────────────────
    def _load_existing_timetable(self):
        nam_hoc_id = self.cmb_nam_hoc.currentData()
        hoc_ky_id  = self.cmb_hoc_ky.currentData()
        if not nam_hoc_id or not hoc_ky_id:
            return
        from core.db.models.thoi_khoa_bieu import ThoiKhoaBieu
        self._current_data.clear()
        for tkb in self.session.query(ThoiKhoaBieu).filter_by(
                nam_hoc_id=nam_hoc_id, hoc_ky_id=hoc_ky_id).all():
            buoi = "Sáng" if tkb.tiet_bat_dau <= 5 else "Chiều"
            tiet = tkb.tiet_bat_dau if buoi == "Sáng" \
                else tkb.tiet_bat_dau - 5
            self._current_data[(tkb.thu, buoi, tiet, tkb.lop_hoc_id)] = \
                (tkb.mon_hoc_id, tkb.giao_vien_id)
        self._render_table()

    # ── Xuất Excel ────────────────────────────────────────────
    def _xuat_excel(self):
        if not self._current_data and not self._row_map:
            QMessageBox.warning(self, "Trống", "Chưa có dữ liệu TKB.")
            return
        from PySide6.QtWidgets import QFileDialog
        path, _ = QFileDialog.getSaveFileName(
            self, "Lưu TKB Excel", "thoi_khoa_bieu.xlsx",
            "Excel (*.xlsx)")
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import (Font, Alignment, PatternFill,
                                          Border, Side)
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Thời khóa biểu"

            visible_lop = self._visible_lop()
            HEADER_FILL = PatternFill("solid", fgColor="1D9E75")
            THU_FILL    = PatternFill("solid", fgColor="E8F4FD")
            SANG_FILL   = PatternFill("solid", fgColor="F0FBF6")
            CHIEU_FILL  = PatternFill("solid", fgColor="FEF9EC")
            thin = Side(style="thin", color="CCCCCC")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            def style(cell, fill=None, bold=False, center=True, wrap=True):
                if fill:
                    cell.fill = fill
                cell.font = Font(bold=bold, size=10)
                cell.alignment = Alignment(
                    horizontal="center" if center else "left",
                    vertical="center", wrap_text=wrap)
                cell.border = border

            # Header hàng 1
            ws.cell(1, 1, "Thứ/Lớp")
            ws.cell(1, 2, "Buổi")
            ws.cell(1, 3, "Tiết")
            for i, lop in enumerate(visible_lop):
                c = ws.cell(1, 4 + i, lop.ten_lop)
                style(c, HEADER_FILL, bold=True)
                ws.column_dimensions[get_column_letter(4 + i)].width = 14
            for col in range(1, 4):
                style(ws.cell(1, col), HEADER_FILL, bold=True)
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 8
            ws.column_dimensions["C"].width = 5

            # Dữ liệu
            excel_row = 2
            for rd in self._row_map:
                thu  = rd["thu"]
                buoi = rd["buoi"]
                tiet = rd["tiet"]
                fill = SANG_FILL if rd["is_sang"] else CHIEU_FILL

                ws.cell(excel_row, 3, tiet)
                style(ws.cell(excel_row, 3), fill)
                ws.row_dimensions[excel_row].height = 32

                for col_i, lop in enumerate(visible_lop):
                    key = (thu, buoi, tiet, lop.id)
                    if key in self._current_data:
                        mon_id, gv_id = self._current_data[key]
                        mon = self._get_mon_by_id(mon_id)
                        gv  = self._get_gv_by_id(gv_id)
                        ma  = mon.ma_mon if mon else "?"
                        ten = ten_tac_gv(
                            gv.nguoi_dung.ho_ten
                            if gv and gv.nguoi_dung else "?")
                        c = ws.cell(excel_row, 4 + col_i, f"{ma}\n{ten}")
                    else:
                        c = ws.cell(excel_row, 4 + col_i, "")
                    style(c)

                excel_row += 1

            # Merge cột Thứ và Buổi
            excel_row = 2
            for thu, day in zip(THU_NUMS, DAYS):
                rows_of_thu = [(i, r) for i, r in enumerate(self._row_map)
                               if r["thu"] == thu]
                if not rows_of_thu:
                    continue
                start_thu = excel_row
                for buoi, is_sang in BUOIS:
                    rows_of_buoi = [(i, r) for i, r in rows_of_thu
                                    if r["buoi"] == buoi]
                    if not rows_of_buoi:
                        continue
                    start_b = excel_row
                    for _, rd in rows_of_buoi:
                        excel_row += 1
                    end_b = excel_row - 1
                    fill = SANG_FILL if is_sang else CHIEU_FILL
                    if start_b < end_b:
                        ws.merge_cells(f"B{start_b}:B{end_b}")
                    c = ws.cell(start_b, 2, buoi)
                    style(c, fill, bold=True)

                end_thu = excel_row - 1
                if start_thu < end_thu:
                    ws.merge_cells(f"A{start_thu}:A{end_thu}")
                c = ws.cell(start_thu, 1, day)
                style(c, THU_FILL, bold=True)

            wb.save(path)
            QMessageBox.information(self, "Thành công",
                                    f"Đã xuất TKB ra:\n{path}")
        except ImportError:
            QMessageBox.critical(self, "Lỗi",
                                 "Cần cài openpyxl: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi xuất Excel", str(e))

    # ── Helpers ───────────────────────────────────────────────
    def _get_mon_by_id(self, mid):
        return next((m for m in self._ds_mon if m.id == mid), None)

    def _get_gv_by_id(self, gid):
        return next((g for g in self._ds_gv if g.id == gid), None)

    def _get_lop_name(self, lop_id):
        lop = next((l for l in self._ds_lop if l.id == lop_id), None)
        return lop.ten_lop if lop else "?"

    def closeEvent(self, event):
        self.session.close()
        super().closeEvent(event)