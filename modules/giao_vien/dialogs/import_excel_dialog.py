# modules/giao_vien/dialogs/import_excel_dialog.py
"""
ImportExcelDialog: nhập danh sách giáo viên từ file Excel.

Luồng:
  1. Chọn file Excel
  2. Preview dữ liệu — highlight lỗi
  3. Xác nhận → import vào DB

Cột Excel chuẩn (xem file mẫu):
  A: Mã GV | B: Họ và tên | C: Email | D: Môn dạy | E: Mã tổ | F: SĐT
"""

import os
from pathlib import Path
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QTableWidget, QTableWidgetItem,
    QFileDialog, QDialogButtonBox, QMessageBox,
    QProgressBar, QHeaderView, QFrame
)
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QColor


# ── Cột Excel ─────────────────────────────────────────────────
EXCEL_COLS = {
    0: ("ma_gv",    "Mã GV",        True),
    1: ("ho_ten",   "Họ và tên",    True),
    2: ("email",    "Email",        True),
    3: ("mon_day",  "Môn dạy",      False),
    4: ("ma_to",    "Mã tổ",        False),
    5: ("so_dt",    "SĐT",          False),
}

COL_COUNT = len(EXCEL_COLS)
REQUIRED_COLS = [k for k, (_, _, req) in EXCEL_COLS.items() if req]

COLOR_OK    = QColor("#F0FBF6")
COLOR_ERROR = QColor("#FDEEEC")
COLOR_WARN  = QColor("#FEF3E6")


class ImportExcelDialog(QDialog):
    def __init__(self, parent=None, ds_to: list = None):
        super().__init__(parent)
        self._ds_to  = {to.ma_to.upper(): to.id for to in (ds_to or [])}
        self._rows   = []      # list[dict] — dữ liệu đã parse
        self._errors = []      # list[str]  — lỗi tổng hợp

        self.setWindowTitle("Nhập giáo viên từ Excel")
        self.setMinimumWidth(760)
        self.setMinimumHeight(520)
        self.setModal(True)
        self._build_ui()

    # ── Build UI ──────────────────────────────────────────────
    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # Hướng dẫn
        note = QLabel(
            "📋  File Excel cần có các cột theo thứ tự: "
            "<b>Mã GV | Họ và tên | Email | Môn dạy | Mã tổ | SĐT</b><br>"
            "Hàng đầu tiên là tiêu đề, dữ liệu bắt đầu từ hàng 2.")
        note.setStyleSheet(
            "background:#E6F1FB; border-radius:6px; padding:10px;"
            " color:#0C447C; font-size:12px;")
        note.setWordWrap(True)
        layout.addWidget(note)

        # Chọn file
        file_row = QHBoxLayout()
        self.lbl_file = QLabel("Chưa chọn file")
        self.lbl_file.setStyleSheet("color:#7A8BAD; font-size:13px;")
        self.lbl_file.setMinimumWidth(360)

        btn_chon = QPushButton("📂  Chọn file Excel")
        btn_chon.setStyleSheet(
            "QPushButton{border:1px solid #DDE2EC;border-radius:6px;"
            "padding:6px 14px;background:#F4F6F9;}"
            "QPushButton:hover{background:#E8EEF7;}")
        btn_chon.clicked.connect(self._on_chon_file)

        self.btn_mau = QPushButton("⬇  Tải file mẫu")
        self.btn_mau.setStyleSheet(
            "QPushButton{border:1px solid #1D9E75;border-radius:6px;"
            "padding:6px 14px;color:#1D9E75;background:white;}"
            "QPushButton:hover{background:#E6F5EF;}")
        self.btn_mau.clicked.connect(self._on_tai_mau)

        file_row.addWidget(self.lbl_file)
        file_row.addStretch()
        file_row.addWidget(self.btn_mau)
        file_row.addWidget(btn_chon)
        layout.addLayout(file_row)

        # Progress
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.progress.setFixedHeight(6)
        self.progress.setStyleSheet(
            "QProgressBar{border:none;background:#E4E8F0;border-radius:3px;}"
            "QProgressBar::chunk{background:#1D9E75;border-radius:3px;}")
        layout.addWidget(self.progress)

        # Label tổng kết
        self.lbl_summary = QLabel()
        self.lbl_summary.setStyleSheet("font-size:12px;")
        self.lbl_summary.setVisible(False)
        layout.addWidget(self.lbl_summary)

        # Bảng preview
        self.table = QTableWidget()
        self.table.setColumnCount(COL_COUNT + 1)  # +1 cột Trạng thái
        headers = [v[1] for v in EXCEL_COLS.values()] + ["Trạng thái"]
        self.table.setHorizontalHeaderLabels(headers)
        self.table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.Stretch)
        self.table.setColumnWidth(0, 80)
        self.table.setColumnWidth(2, 160)
        self.table.setColumnWidth(4, 80)
        self.table.setColumnWidth(5, 100)
        self.table.setColumnWidth(6, 120)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setAlternatingRowColors(False)
        self.table.verticalHeader().setVisible(False)
        layout.addWidget(self.table)

        # Buttons
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("background:#E4E8F0; max-height:1px;")
        layout.addWidget(sep)

        self.btns = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.btns.button(QDialogButtonBox.Ok).setText("✓  Nhập vào hệ thống")
        self.btns.button(QDialogButtonBox.Ok).setEnabled(False)
        self.btns.button(QDialogButtonBox.Ok).setStyleSheet(
            "QPushButton:enabled{background:#1D9E75;color:white;"
            "border-radius:6px;padding:6px 18px;font-weight:600;border:none;}"
            "QPushButton:disabled{background:#B2D8CC;color:white;"
            "border-radius:6px;padding:6px 18px;border:none;}")
        self.btns.button(QDialogButtonBox.Cancel).setText("Huỷ")
        self.btns.accepted.connect(self._on_import)
        self.btns.rejected.connect(self.reject)
        layout.addWidget(self.btns)

    # ── Chọn file ─────────────────────────────────────────────
    def _on_chon_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Chọn file Excel", "",
            "Excel Files (*.xlsx *.xls *.csv)")
        if not path:
            return
        self.lbl_file.setText(Path(path).name)
        self.lbl_file.setStyleSheet("color:#1A1D23; font-size:13px;")
        self._parse_file(path)

    def _parse_file(self, path: str):
        try:
            import pandas as pd
        except ImportError:
            QMessageBox.critical(self, "Lỗi",
                                 "Cần cài pandas: pip install pandas openpyxl")
            return

        self.progress.setVisible(True)
        self.progress.setValue(30)

        try:
            if path.endswith(".csv"):
                df = pd.read_csv(path, dtype=str, encoding="utf-8-sig")
            else:
                df = pd.read_excel(path, dtype=str, header=0)

            self.progress.setValue(70)
            df = df.fillna("")

            self._rows = []
            self._errors = []

            for idx, row in df.iterrows():
                values = list(row.values)
                # Padding nếu thiếu cột
                while len(values) < COL_COUNT:
                    values.append("")

                parsed = {}
                row_errors = []

                for col_idx, (field, label, required) in EXCEL_COLS.items():
                    val = str(values[col_idx]).strip() if col_idx < len(values) else ""
                    parsed[field] = val
                    if required and not val:
                        row_errors.append(f"Thiếu {label}")

                # Validate email
                email = parsed.get("email", "")
                if email and "@" not in email:
                    row_errors.append("Email không hợp lệ")

                # Kiểm tra mã tổ
                ma_to = parsed.get("ma_to", "").upper()
                if ma_to and ma_to not in self._ds_to:
                    row_errors.append(f"Mã tổ '{ma_to}' không tồn tại")
                elif ma_to:
                    parsed["to_id"] = self._ds_to[ma_to]
                else:
                    parsed["to_id"] = None

                parsed["_row_errors"] = row_errors
                parsed["_excel_row"]  = idx + 2  # +2 vì header ở row 1
                self._rows.append(parsed)

            self.progress.setValue(100)
            self._render_preview()

        except Exception as e:
            QMessageBox.critical(self, "Lỗi đọc file", str(e))
        finally:
            self.progress.setVisible(False)

    # ── Render preview ────────────────────────────────────────
    def _render_preview(self):
        self.table.setRowCount(0)
        error_count = 0
        ok_count    = 0

        for i, row in enumerate(self._rows):
            self.table.insertRow(i)
            has_error = bool(row["_row_errors"])
            bg = COLOR_ERROR if has_error else COLOR_OK

            for col_idx, (field, _, _) in EXCEL_COLS.items():
                val  = row.get(field, "")
                item = QTableWidgetItem(val)
                item.setBackground(bg)
                self.table.setItem(i, col_idx, item)

            # Cột trạng thái
            if has_error:
                status = "⚠  " + "; ".join(row["_row_errors"])
                st_item = QTableWidgetItem(status)
                st_item.setForeground(QColor("#993C1D"))
                error_count += 1
            else:
                st_item = QTableWidgetItem("✓ Hợp lệ")
                st_item.setForeground(QColor("#0F6E56"))
                ok_count += 1

            st_item.setBackground(bg)
            self.table.setItem(i, COL_COUNT, st_item)

        # Tổng kết
        total = len(self._rows)
        if error_count == 0:
            summary = (f"✓  {total} dòng — tất cả hợp lệ, sẵn sàng nhập.")
            self.lbl_summary.setStyleSheet("color:#0F6E56; font-weight:600;")
        else:
            summary = (f"⚠  {total} dòng: {ok_count} hợp lệ, "
                       f"{error_count} lỗi (sẽ bỏ qua khi nhập).")
            self.lbl_summary.setStyleSheet("color:#854F0B; font-weight:600;")

        self.lbl_summary.setText(summary)
        self.lbl_summary.setVisible(True)
        self.btns.button(QDialogButtonBox.Ok).setEnabled(ok_count > 0)

    # ── Tải file mẫu ──────────────────────────────────────────
    def _on_tai_mau(self):
        save_path, _ = QFileDialog.getSaveFileName(
            self, "Lưu file mẫu", "mau_giao_vien.xlsx",
            "Excel Files (*.xlsx)")
        if not save_path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Giáo viên"

            # Header
            headers = [v[1] for v in EXCEL_COLS.values()]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = PatternFill("solid", fgColor="1D9E75")
                cell.alignment = Alignment(horizontal="center")

            # Dữ liệu mẫu
            sample = [
                ["GV001", "Nguyễn Văn A", "nva@truong.edu.vn",
                 "Toán",   "TO_TOAN",  "0901234567"],
                ["GV002", "Trần Thị B",   "ttb@truong.edu.vn",
                 "Vật lý", "KHOA_HOC", "0912345678"],
                ["GV003", "Lê Văn C",     "lvc@truong.edu.vn",
                 "Hoá học","KHOA_HOC", ""],
            ]
            for row_data in sample:
                ws.append(row_data)

            # Độ rộng cột
            widths = [10, 22, 28, 12, 12, 14]
            for col, w in enumerate(widths, 1):
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col)].width = w

            wb.save(save_path)
            QMessageBox.information(
                self, "Thành công",
                f"Đã lưu file mẫu:\n{save_path}")
        except ImportError:
            QMessageBox.critical(self, "Lỗi",
                                 "Cần cài openpyxl: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    # ── Import ────────────────────────────────────────────────
    def _on_import(self):
        valid_rows = [r for r in self._rows if not r["_row_errors"]]
        if not valid_rows:
            QMessageBox.warning(self, "Không có dữ liệu hợp lệ",
                                "Không có dòng nào hợp lệ để nhập.")
            return
        self.accept()

    # ── Lấy dữ liệu hợp lệ ───────────────────────────────────
    def get_valid_rows(self) -> list[dict]:
        """
        Trả về list các dict hợp lệ để service xử lý.
        Mỗi dict gồm: ma_gv, ho_ten, email, mon_day, to_id, so_dt
        """
        return [
            {
                "ma_gv":   r["ma_gv"],
                "ho_ten":  r["ho_ten"],
                "email":   r["email"],
                "mon_day": r["mon_day"],
                "to_id":   r.get("to_id"),
                "so_dt":   r["so_dt"],
            }
            for r in self._rows
            if not r["_row_errors"]
        ]
