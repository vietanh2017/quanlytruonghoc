# modules/reports/service/export_excel_service.py

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins, PrintOptions
from shared.dto.result import ServiceResult
from modules.reports.service.bao_cao_service import BaoCaoService


class ExportExcelService:
    def __init__(self):
        self.svc = BaoCaoService()
        self.wb = None

    def export_report(self, loai: str, nam_hoc_id: int, ky: int, 
                     thang: int, chon_thang: bool, export_path: str, options: dict):
        """Xuất báo cáo Excel"""
        try:
            self.wb = Workbook()
            self.wb.remove(self.wb.active)

            if loai == "giao_vien":
                self._export_giao_vien(nam_hoc_id, ky, thang, chon_thang, options)
            else:
                self._export_hoc_sinh(nam_hoc_id, ky, thang, chon_thang, options)

            self.wb.save(export_path)
            return ServiceResult(ok=True, error="Xuất thành công")

        except Exception as e:
            return ServiceResult(ok=False, error=str(e))

    def _setup_sheet(self, ws, title, headers, col_widths=None):
        """Thiết lập sheet (không khóa)"""
        # === 1. TIÊU ĐỀ ===
        ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        cell = ws['A1']
        cell.value = title
        cell.font = Font(size=16, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="1D9E75", end_color="1D9E75", fill_type="solid")
        ws.row_dimensions[1].height = 35

        # === 2. HEADER BẢNG ===
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="0F6E56", end_color="0F6E56", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin', color='FFFFFF'),
                right=Side(style='thin', color='FFFFFF'),
                top=Side(style='thin', color='FFFFFF'),
                bottom=Side(style='thin', color='FFFFFF')
            )
        ws.row_dimensions[2].height = 28

        # === 3. FREEZE PANES ===
        ws.freeze_panes = 'A3'

        # === 4. IN ẤN ===
        ws.page_margins = PageMargins(left=0.7, right=0.7, top=0.7, bottom=0.7)
        ws.print_options = PrintOptions(horizontalCentered=True, verticalCentered=False)
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.paperSize = 9
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        # === 5. ĐỘ RỘNG CỘT ===
        if col_widths:
            for col, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
        else:
            # Mặc định nếu không truyền
            default_widths = [8, 18, 25, 15, 18, 15, 12]  # STT, Mã, Tên, Điểm, Xếp loại, ...
            for col, width in enumerate(default_widths[:len(headers)], 1):
                ws.column_dimensions[get_column_letter(col)].width = width

        return 3

    def _apply_cell_style(self, cell, bold=False, align='center'):
        """Áp dụng style cho ô dữ liệu"""
        cell.font = Font(size=10, bold=bold)
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

    def _export_giao_vien(self, nam_hoc_id, ky, thang, chon_thang, options):
        """Xuất báo cáo giáo viên"""
        
        if chon_thang and thang is None:
            for t in [9, 10, 11, 12, 1, 2, 3, 4, 5]:
                self._create_gv_thang_sheet(nam_hoc_id, t)
        else:
            self._create_gv_thang_sheet(nam_hoc_id, thang)

        self._create_gv_hk_sheet(nam_hoc_id, 1, "Học kỳ I")
        self._create_gv_hk_sheet(nam_hoc_id, 2, "Học kỳ II")
        self._create_gv_ca_nam_sheet(nam_hoc_id)

    def _create_gv_thang_sheet(self, nam_hoc_id, thang):
        sheet_name = f"Tháng {thang}" if thang else "Tất cả tháng"
        ws = self.wb.create_sheet(sheet_name)

        headers = ["STT", "Mã GV", "Họ tên", "Điểm", "Xếp loại"]
        col_widths = [8, 15, 30, 15, 25]  # ← Cột Họ tên rộng 30
        start_row = self._setup_sheet(ws, f"BÁO CÁO THI ĐUA GIÁO VIÊN - {sheet_name.upper()}", headers, col_widths)

        result = self.svc.xep_hang_giao_vien(nam_hoc_id, None, thang)
        if result.ok:
            data = result.data
            for row, item in enumerate(data, start_row):
                ws.cell(row=row, column=1, value=row-2)
                ws.cell(row=row, column=2, value=item['ma'])
                ws.cell(row=row, column=3, value=item['ten'])
                ws.cell(row=row, column=4, value=item['diem'])
                ws.cell(row=row, column=5, value=item['xep_loai'])
                
                for col in range(1, 6):
                    self._apply_cell_style(ws.cell(row=row, column=col))

    def _create_gv_hk_sheet(self, nam_hoc_id, hoc_ky_id, ten_hk):
        """Tạo sheet học kỳ"""
        ws = self.wb.create_sheet(ten_hk)

        headers = ["STT", "Mã GV", "Họ tên", "Điểm", "Xếp loại"]
        start_row = self._setup_sheet(ws, f"BÁO CÁO THI ĐUA GIÁO VIÊN - {ten_hk.upper()}", headers)

        result = self.svc.xep_hang_giao_vien(nam_hoc_id, hoc_ky_id, None)
        if result.ok:
            data = result.data
            for row, item in enumerate(data, start_row):
                ws.cell(row=row, column=1, value=row-2)
                ws.cell(row=row, column=2, value=item['ma'])
                ws.cell(row=row, column=3, value=item['ten'])
                ws.cell(row=row, column=4, value=item['diem'])
                ws.cell(row=row, column=5, value=item['xep_loai'])
                
                for col in range(1, 6):
                    self._apply_cell_style(ws.cell(row=row, column=col))

    def _create_gv_ca_nam_sheet(self, nam_hoc_id):
        """Tạo sheet cả năm"""
        ws = self.wb.create_sheet("Cả năm")

        headers = ["STT", "Mã GV", "Họ tên", "Điểm", "Xếp loại"]
        start_row = self._setup_sheet(ws, "BÁO CÁO THI ĐUA GIÁO VIÊN - CẢ NĂM", headers)

        result = self.svc.xep_hang_giao_vien(nam_hoc_id, None, None)
        if result.ok:
            data = result.data
            for row, item in enumerate(data, start_row):
                ws.cell(row=row, column=1, value=row-2)
                ws.cell(row=row, column=2, value=item['ma'])
                ws.cell(row=row, column=3, value=item['ten'])
                ws.cell(row=row, column=4, value=item['diem'])
                ws.cell(row=row, column=5, value=item['xep_loai'])
                
                for col in range(1, 6):
                    self._apply_cell_style(ws.cell(row=row, column=col))

    def _export_hoc_sinh(self, nam_hoc_id, ky, thang, chon_thang, options):
        """Xuất báo cáo học sinh"""
        sheet_name = "Học sinh"
        ws = self.wb.create_sheet(sheet_name)
        
        headers = ["STT", "Lớp", "Điểm TB học tập", "Điểm TB Đội", "Điểm TB", "Xếp thứ"]
        start_row = self._setup_sheet(ws, "BÁO CÁO THI ĐUA HỌC SINH", headers)

        result = self.svc.xep_hang_lop(nam_hoc_id, None, thang)
        if result.ok:
            data = result.data
            for row, item in enumerate(data, start_row):
                ws.cell(row=row, column=1, value=row-2)
                ws.cell(row=row, column=2, value=item['ten'])
                # Cần thêm điểm chi tiết
                ws.cell(row=row, column=3, value=0)
                ws.cell(row=row, column=4, value=0)
                ws.cell(row=row, column=5, value=item['diem'])
                ws.cell(row=row, column=6, value=item['thu_hang'])
                
                for col in range(1, 7):
                    self._apply_cell_style(ws.cell(row=row, column=col))